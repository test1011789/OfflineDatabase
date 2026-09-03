from __future__ import annotations

import json
import shutil
import sqlite3
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

import tkinter as tk
from tkinter import colorchooser, filedialog, messagebox, simpledialog, ttk

from openpyxl import load_workbook


APP_TITLE = '離線資料庫管理工具'
DEFAULT_XLSX = 'initial_data.xlsx'
LINK_MARKER = '📎'
DEFAULT_LINK_COLUMN_WIDTH = 24
PRODUCTION_MARKER = '🏭'
DEFAULT_PRODUCTION_COLUMN_WIDTH = 34


DEFAULT_LOOKUPS = [
    ('EC', '乳劑'),
    ('SC', '水懸劑'),
    ('SG', '水溶性粒劑'),
    ('SL', '溶液'),
    ('SP', '水溶性粉劑'),
    ('WG', '水分散性粒劑'),
    ('WP', '可溼性粉劑'),
    ('UL', '超低容量液劑'),
    ('EW', '水基乳劑'),
]

DEFAULT_CATEGORY_COLORS = {
    '除草劑': '#E2F0D9',
    '殺蟲劑': '#FCE4D6',
    '殺菌劑': '#DDEBF7',
}

DEFAULT_COLUMN_WIDTH = 140
FIXED_UI_FONT_SIZE = 10

BUILTIN_FIELD_KEYS = {
    '核准': 'approval',
    '生產': 'production',
    '許可證號碼\nReg. NO.': 'permit_number',
    '劑型': 'formulation',
    '含量\nA.I.': 'content',
    '劑型種類\nTerm': 'formulation_term',
    '劑型種類代碼\nCode': 'formulation_code',
    '中文普通名稱\nCommon name': 'common_name',
    '成品名\nBrand Name': 'brand_name',
    '倉庫白板貼': 'warehouse_board',
    '作物': 'crop',
    '使用特殊瓶': 'special_bottle',
    '雅飛總經銷': 'distributor',
    '版本': 'version',
    '委外加工': 'outsourcing',
}

DEFAULT_DISPLAY_SETTINGS = {
    'header_font_size': 10,
    'data_font_size': 10,
    'header_row_height': 42,
    'data_row_height': 28,
    'column_widths': {},
    'link_column_widths': {},
    'production_column_widths': {},
    'show_link_marker': True,
    'link_marker': LINK_MARKER,
    'production_visible_columns': ['date', 'order', 'stock', 'manufacturer', 'remark'],
}


# -----------------------------
# Portable paths and database
# -----------------------------
def app_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def copy_file_portable(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with Path(source).open('rb') as src, Path(destination).open('wb') as dst:
        while True:
            chunk = src.read(1024 * 1024)
            if not chunk:
                break
            dst.write(chunk)


class Database:
    def __init__(self, root: Path):
        self.root = Path(root)
        self.data_dir = self.root / 'data'
        self.backup_dir = self.root / 'backup'
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        self.path = self.data_dir / 'app.db'
        self.conn = sqlite3.connect(self.path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute('PRAGMA foreign_keys = ON')
        self._init_schema()

    def _init_schema(self) -> None:
        self.conn.executescript(
            '''
            CREATE TABLE IF NOT EXISTS app_meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                position INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS fields (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                kind TEXT NOT NULL DEFAULT 'text',
                options_json TEXT NOT NULL DEFAULT '[]',
                color_map_json TEXT NOT NULL DEFAULT '{}',
                position INTEGER NOT NULL,
                system_key TEXT,
                builtin INTEGER NOT NULL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                company_id INTEGER NOT NULL DEFAULT 1,
                FOREIGN KEY (company_id) REFERENCES companies(id)
            );
            CREATE TABLE IF NOT EXISTS record_values (
                record_id INTEGER NOT NULL,
                field_id INTEGER NOT NULL,
                value TEXT,
                PRIMARY KEY (record_id, field_id),
                FOREIGN KEY (record_id) REFERENCES records(id) ON DELETE CASCADE,
                FOREIGN KEY (field_id) REFERENCES fields(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS record_links (
                record_id INTEGER NOT NULL,
                field_id INTEGER NOT NULL,
                url TEXT NOT NULL,
                PRIMARY KEY (record_id, field_id),
                FOREIGN KEY (record_id) REFERENCES records(id) ON DELETE CASCADE,
                FOREIGN KEY (field_id) REFERENCES fields(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS production_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_id INTEGER NOT NULL,
                production_date TEXT NOT NULL DEFAULT '',
                batch_no TEXT NOT NULL DEFAULT '',
                quantity TEXT NOT NULL DEFAULT '',
                order_quantity TEXT NOT NULL DEFAULT '',
                stock_quantity TEXT NOT NULL DEFAULT '',
                manufacturer TEXT NOT NULL DEFAULT '',
                remark TEXT NOT NULL DEFAULT '',
                external_url TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (record_id) REFERENCES records(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_production_records_record_id ON production_records(record_id);
            CREATE INDEX IF NOT EXISTS idx_production_records_record_date ON production_records(record_id, production_date DESC, id DESC);
            CREATE INDEX IF NOT EXISTS idx_production_records_manufacturer ON production_records(manufacturer);
            CREATE TABLE IF NOT EXISTS production_manufacturers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                company_id INTEGER NOT NULL DEFAULT 1,
                position INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY (company_id) REFERENCES companies(id),
                UNIQUE(company_id, name)
            );
            CREATE TABLE IF NOT EXISTS lookup_pairs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                term TEXT NOT NULL,
                position INTEGER NOT NULL DEFAULT 0
            );
            '''
        )
        # 公司資料層：舊版資料全部歸入第一家公司，並建立第二家公司。
        company_rows = list(self.conn.execute('SELECT id, name FROM companies ORDER BY position, id'))
        if not company_rows:
            self.conn.execute("INSERT INTO companies(name, position) VALUES ('公司 A', 0)")
            self.conn.execute("INSERT INTO companies(name, position) VALUES ('公司 B', 1)")
        else:
            names = {str(r['name']) for r in company_rows}
            if '公司 A' not in names:
                self.conn.execute("INSERT INTO companies(name, position) VALUES ('公司 A', COALESCE((SELECT MAX(position)+1 FROM companies), 0))")
            if '公司 B' not in names:
                self.conn.execute("INSERT INTO companies(name, position) VALUES ('公司 B', COALESCE((SELECT MAX(position)+1 FROM companies), 0))")
        company_ids = [int(r['id']) for r in self.conn.execute('SELECT id FROM companies ORDER BY position, id')]
        default_company_id = company_ids[0]

        record_columns = {row['name'] for row in self.conn.execute('PRAGMA table_info(records)')}
        if 'company_id' not in record_columns:
            self.conn.execute('ALTER TABLE records ADD COLUMN company_id INTEGER NOT NULL DEFAULT 1')
        self.conn.execute('UPDATE records SET company_id = ? WHERE company_id IS NULL OR company_id <= 0', (default_company_id,))
        self.conn.execute('CREATE INDEX IF NOT EXISTS idx_records_company_id ON records(company_id)')

        manufacturer_columns = {row['name'] for row in self.conn.execute('PRAGMA table_info(production_manufacturers)')}
        if 'company_id' not in manufacturer_columns:
            # 舊版 manufacturer 表的 name 全域唯一，直接加入公司欄位即可；同名廠商仍可在兩家公司各自使用。
            self.conn.execute('ALTER TABLE production_manufacturers ADD COLUMN company_id INTEGER NOT NULL DEFAULT 1')
        self.conn.execute('UPDATE production_manufacturers SET company_id = ? WHERE company_id IS NULL OR company_id <= 0', (default_company_id,))
        # V5 的廠商表是 name 全域唯一。V6+ 需要改成「每家公司各自唯一」，因此第一次升級時重建此表。
        unique_name_index = False
        for idx in self.conn.execute('PRAGMA index_list(production_manufacturers)').fetchall():
            if int(idx['unique']) != 1:
                continue
            idx_name = str(idx['name'])
            cols = [str(r['name']) for r in self.conn.execute(f'PRAGMA index_info("{idx_name.replace(chr(34), chr(34)*2)}")').fetchall()]
            if cols == ['name']:
                unique_name_index = True
                break
        if unique_name_index:
            self.conn.execute('ALTER TABLE production_manufacturers RENAME TO production_manufacturers_old')
            self.conn.execute('''
                CREATE TABLE production_manufacturers (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    company_id INTEGER NOT NULL DEFAULT 1,
                    position INTEGER NOT NULL DEFAULT 0,
                    FOREIGN KEY (company_id) REFERENCES companies(id),
                    UNIQUE(company_id, name)
                )
            ''')
            self.conn.execute('''
                INSERT INTO production_manufacturers(id, name, company_id, position)
                SELECT id, name, company_id, position FROM production_manufacturers_old
            ''')
            self.conn.execute('DROP TABLE production_manufacturers_old')
        self.conn.execute('CREATE INDEX IF NOT EXISTS idx_production_manufacturers_company ON production_manufacturers(company_id, position, id)')
        self.conn.execute('CREATE INDEX IF NOT EXISTS idx_records_company_id_id ON records(company_id, id)')

        production_columns = {row['name'] for row in self.conn.execute('PRAGMA table_info(production_records)')}
        if 'order_quantity' not in production_columns:
            self.conn.execute("ALTER TABLE production_records ADD COLUMN order_quantity TEXT NOT NULL DEFAULT ''")
        if 'stock_quantity' not in production_columns:
            self.conn.execute("ALTER TABLE production_records ADD COLUMN stock_quantity TEXT NOT NULL DEFAULT ''")
        if 'manufacturer' not in production_columns:
            self.conn.execute("ALTER TABLE production_records ADD COLUMN manufacturer TEXT NOT NULL DEFAULT ''")
        if 'external_url' not in production_columns:
            self.conn.execute("ALTER TABLE production_records ADD COLUMN external_url TEXT NOT NULL DEFAULT ''")
        # 舊版的「quantity」資料視為下單數量，讓既有資料不會消失。
        self.conn.execute("UPDATE production_records SET order_quantity = quantity WHERE TRIM(order_quantity) = '' AND TRIM(quantity) <> ''")
        existing_manufacturers = [
            str(row['manufacturer'] or '').strip()
            for row in self.conn.execute("SELECT DISTINCT manufacturer FROM production_records WHERE TRIM(manufacturer) <> '' ORDER BY manufacturer")
        ]
        for name in existing_manufacturers:
            self.conn.execute(
                'INSERT OR IGNORE INTO production_manufacturers(name, position) VALUES (?, COALESCE((SELECT MAX(position) + 1 FROM production_manufacturers), 0))',
                (name,),
            )

        field_columns = {row['name'] for row in self.conn.execute('PRAGMA table_info(fields)')}
        if 'system_key' not in field_columns:
            self.conn.execute('ALTER TABLE fields ADD COLUMN system_key TEXT')
        self._backfill_system_keys()
        for code, term in DEFAULT_LOOKUPS:
            self.conn.execute(
                'INSERT OR IGNORE INTO lookup_pairs(code, term, position) VALUES (?, ?, ?)',
                (code, term, len(DEFAULT_LOOKUPS)),
            )
        self.conn.commit()

    def _backfill_system_keys(self) -> None:
        for name, system_key in BUILTIN_FIELD_KEYS.items():
            self.conn.execute(
                'UPDATE fields SET system_key = ? WHERE system_key IS NULL AND name = ?',
                (system_key, name),
            )
        self.conn.commit()

    @staticmethod
    def builtin_key_for_name(name: str) -> str | None:
        return BUILTIN_FIELD_KEYS.get(name)

    def close(self) -> None:
        self.conn.close()

    def has_data(self) -> bool:
        return self.conn.execute('SELECT COUNT(*) FROM fields').fetchone()[0] > 0

    def import_excel(self, xlsx_path: Path, replace: bool = False, company_id: int | None = None) -> tuple[int, int]:
        company_id = int(company_id or self.default_company_id())
        wb_values = load_workbook(xlsx_path, data_only=True, read_only=False)
        wb_formulas = load_workbook(xlsx_path, data_only=False, read_only=False)
        ws = wb_values[wb_values.sheetnames[0]]
        ws_formula = wb_formulas[wb_formulas.sheetnames[0]]

        headers: list[str] = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(1, col).value
            header = str(value).strip() if value is not None else f'欄位{col}'
            headers.append(header)

        descriptions_row = None
        for row in range(2, ws.max_row + 1):
            first_value = ws.cell(row, 1).value
            if isinstance(first_value, str) and first_value.strip() == '欄位說明':
                descriptions_row = row
                break

        end_row = (descriptions_row - 1) if descriptions_row else ws.max_row
        data_rows: list[int] = []
        for row in range(2, end_row + 1):
            values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
            if any(value not in (None, '') for value in values):
                data_rows.append(row)

        if replace:
            self.conn.execute('DELETE FROM production_records WHERE record_id IN (SELECT id FROM records WHERE company_id = ?)', (company_id,))
            self.conn.execute('DELETE FROM record_links WHERE record_id IN (SELECT id FROM records WHERE company_id = ?)', (company_id,))
            self.conn.execute('DELETE FROM record_values WHERE record_id IN (SELECT id FROM records WHERE company_id = ?)', (company_id,))
            self.conn.execute('DELETE FROM records WHERE company_id = ?', (company_id,))

        current_fields = self.list_fields()
        existing_names = {row['name']: row for row in current_fields}
        existing_keys = {row['system_key']: row for row in current_fields if row['system_key']}
        field_ids: list[int] = []
        for position, name in enumerate(headers):
            system_key = self.builtin_key_for_name(name)
            existing = existing_keys.get(system_key) if system_key else existing_names.get(name)
            if existing:
                field_id = int(existing['id'])
                if system_key and existing['system_key'] != system_key:
                    self.conn.execute('UPDATE fields SET system_key = ? WHERE id = ?', (system_key, field_id))
            else:
                kind, options, colors = self._infer_field(name)
                cursor = self.conn.execute(
                    '''INSERT INTO fields(name, kind, options_json, color_map_json, position, system_key, builtin)
                       VALUES (?, ?, ?, ?, ?, ?, 1)''',
                    (name, kind, json.dumps(options, ensure_ascii=False), json.dumps(colors, ensure_ascii=False), position, system_key),
                )
                field_id = cursor.lastrowid
            field_ids.append(int(field_id))

        lookup_by_term = {term: code for code, term in DEFAULT_LOOKUPS}
        for row in data_rows:
            record_id = self.create_record({}, company_id=company_id)
            for col, field_id in enumerate(field_ids, start=1):
                value = ws.cell(row, col).value
                formula_cell = ws_formula.cell(row, col).value
                # Array formulas in the source workbook are represented as objects by openpyxl.
                if col == 7 and (value is None or not isinstance(value, (str, int, float))):
                    value = lookup_by_term.get(str(ws.cell(row, 6).value).strip(), '')
                if value is None:
                    value = ''
                if isinstance(value, float) and value.is_integer():
                    value = str(int(value))
                self.set_value(record_id, field_id, str(value))
                hyperlink = ws.cell(row, col).hyperlink
                if hyperlink and hyperlink.target:
                    self.conn.execute(
                        'INSERT OR REPLACE INTO record_links(record_id, field_id, url) VALUES (?, ?, ?)',
                        (record_id, field_id, hyperlink.target),
                    )
            self._sync_linked_term(record_id)

        self.conn.execute(
            'INSERT OR REPLACE INTO app_meta(key, value) VALUES (?, ?)',
            ('source_file', str(xlsx_path)),
        )
        self.conn.execute(
            'INSERT OR REPLACE INTO app_meta(key, value) VALUES (?, ?)',
            ('imported_at', datetime.now().isoformat(timespec='seconds')),
        )
        self.conn.commit()
        return len(headers), len(data_rows)

    @staticmethod
    def _infer_field(name: str) -> tuple[str, list[str], dict[str, str]]:
        if name in {'核准', '生產', '倉庫白板貼', '雅飛總經銷'}:
            return 'dropdown', ['O', 'X'], {}
        if name == '劑型':
            return 'category', ['除草劑', '殺蟲劑', '殺菌劑'], DEFAULT_CATEGORY_COLORS.copy()
        if name == '劑型種類\nTerm':
            return 'computed', [], {}
        if name == '劑型種類代碼\nCode':
            return 'lookup', [], {}
        if name == '版本':
            return 'dropdown', ['1', '2', '3'], {}
        return 'text', [], {}

    def get_display_settings(self) -> dict[str, Any]:
        settings = dict(DEFAULT_DISPLAY_SETTINGS)
        row = self.conn.execute('SELECT value FROM app_meta WHERE key = ?', ('display_settings',)).fetchone()
        if row:
            try:
                stored = json.loads(row['value'])
                if isinstance(stored, dict):
                    # 相容前一版設定：舊的 row_height 先沿用為資料列高。
                    if 'data_row_height' not in stored and 'row_height' in stored:
                        stored['data_row_height'] = stored['row_height']
                    # 相容前一版單一字體設定，先同時套用到標題與資料。
                    if 'header_font_size' not in stored and 'font_size' in stored:
                        stored['header_font_size'] = stored['font_size']
                    if 'data_font_size' not in stored and 'font_size' in stored:
                        stored['data_font_size'] = stored['font_size']
                    settings.update(stored)
            except (TypeError, ValueError, json.JSONDecodeError):
                pass
        settings.pop('font_size', None)
        settings.pop('row_height', None)
        settings.pop('default_column_width', None)
        settings.pop('form_gap', None)
        if not isinstance(settings.get('column_widths'), dict):
            settings['column_widths'] = {}
        if not isinstance(settings.get('link_column_widths'), dict):
            settings['link_column_widths'] = {}
        if not isinstance(settings.get('production_column_widths'), dict):
            settings['production_column_widths'] = {}
        if not isinstance(settings.get('show_link_marker'), bool):
            settings['show_link_marker'] = True
        if not isinstance(settings.get('link_marker'), str):
            settings['link_marker'] = LINK_MARKER
        allowed_production_columns = {'date', 'order', 'stock', 'manufacturer', 'remark'}
        visible = settings.get('production_visible_columns')
        if not isinstance(visible, list):
            settings['production_visible_columns'] = ['date', 'order', 'stock', 'manufacturer', 'remark']
        else:
            settings['production_visible_columns'] = [str(v) for v in visible if str(v) in allowed_production_columns]
            if not settings['production_visible_columns']:
                settings['production_visible_columns'] = ['date', 'order', 'stock', 'manufacturer', 'remark']
        return settings

    def save_display_settings(self, settings: dict[str, Any]) -> None:
        merged = dict(DEFAULT_DISPLAY_SETTINGS)
        merged.update(settings)
        merged.pop('font_size', None)
        merged.pop('row_height', None)
        merged.pop('default_column_width', None)
        merged.pop('form_gap', None)
        self.conn.execute(
            'INSERT OR REPLACE INTO app_meta(key, value) VALUES (?, ?)',
            ('display_settings', json.dumps(merged, ensure_ascii=False)),
        )
        self.conn.commit()

    def list_fields(self, active_only: bool = True) -> list[sqlite3.Row]:
        sql = 'SELECT * FROM fields'
        if active_only:
            sql += ' WHERE active = 1'
        sql += ' ORDER BY position, id'
        return list(self.conn.execute(sql))

    def get_field(self, field_id: int) -> sqlite3.Row | None:
        return self.conn.execute('SELECT * FROM fields WHERE id = ?', (field_id,)).fetchone()

    def reorder_fields(self, ordered_ids: list[int]) -> None:
        for position, field_id in enumerate(ordered_ids):
            self.conn.execute('UPDATE fields SET position = ? WHERE id = ?', (position, int(field_id)))
        self.conn.commit()

    def rename_field(self, field_id: int, name: str) -> None:
        name = name.strip()
        if not name:
            raise ValueError('欄位名稱不可空白。')
        duplicate = self.conn.execute(
            'SELECT 1 FROM fields WHERE name = ? AND id <> ?', (name, field_id)
        ).fetchone()
        if duplicate:
            raise ValueError('欄位名稱已存在。')
        self.conn.execute('UPDATE fields SET name = ? WHERE id = ?', (name, field_id))
        self.conn.commit()

    def add_field(self, name: str, kind: str, options: list[str]) -> int:
        name = name.strip()
        if not name:
            raise ValueError('欄位名稱不可空白。')
        if self.conn.execute('SELECT 1 FROM fields WHERE name = ?', (name,)).fetchone():
            raise ValueError('欄位名稱已存在。')
        position = self.conn.execute('SELECT COALESCE(MAX(position), -1) + 1 FROM fields').fetchone()[0]
        cursor = self.conn.execute(
            '''INSERT INTO fields(name, kind, options_json, color_map_json, position, builtin)
               VALUES (?, ?, ?, '{}', ?, 0)''',
            (name, kind, json.dumps(options, ensure_ascii=False), position),
        )
        self.conn.commit()
        return int(cursor.lastrowid)

    def update_field_options(self, field_id: int, options: list[str]) -> None:
        self.conn.execute(
            'UPDATE fields SET options_json = ? WHERE id = ?',
            (json.dumps(options, ensure_ascii=False), field_id),
        )
        self.conn.commit()

    def update_field_colors(self, field_id: int, colors: dict[str, str]) -> None:
        self.conn.execute(
            'UPDATE fields SET color_map_json = ? WHERE id = ?',
            (json.dumps(colors, ensure_ascii=False), field_id),
        )
        self.conn.commit()

    def companies(self) -> list[sqlite3.Row]:
        return list(self.conn.execute('SELECT id, name, position FROM companies ORDER BY position, id'))

    def default_company_id(self) -> int:
        row = self.conn.execute('SELECT id FROM companies ORDER BY position, id LIMIT 1').fetchone()
        if not row:
            self.conn.execute("INSERT INTO companies(name, position) VALUES ('公司 A', 0)")
            self.conn.execute("INSERT INTO companies(name, position) VALUES ('公司 B', 1)")
            self.conn.commit()
            row = self.conn.execute('SELECT id FROM companies ORDER BY position, id LIMIT 1').fetchone()
        return int(row['id'])

    def rename_company(self, company_id: int, name: str) -> None:
        name = name.strip()
        if not name:
            raise ValueError('公司名稱不可空白。')
        if self.conn.execute('SELECT 1 FROM companies WHERE name = ? AND id <> ?', (name, int(company_id))).fetchone():
            raise ValueError('公司名稱已存在。')
        self.conn.execute('UPDATE companies SET name = ? WHERE id = ?', (name, int(company_id)))
        self.conn.commit()

    def add_company(self, name: str) -> int:
        name = name.strip()
        if not name:
            raise ValueError('公司名稱不可空白。')
        if self.conn.execute('SELECT 1 FROM companies WHERE name = ?', (name,)).fetchone():
            raise ValueError('公司名稱已存在。')
        row = self.conn.execute('SELECT COALESCE(MAX(position), -1) + 1 AS next_position FROM companies').fetchone()
        position = int(row['next_position'])
        cursor = self.conn.execute('INSERT INTO companies(name, position) VALUES (?, ?)', (name, position))
        self.conn.commit()
        return int(cursor.lastrowid)

    def delete_company(self, company_id: int) -> None:
        company_id = int(company_id)
        company = self.conn.execute('SELECT name FROM companies WHERE id = ?', (company_id,)).fetchone()
        if not company:
            raise ValueError('找不到要刪除的公司。')
        count = self.conn.execute('SELECT COUNT(*) AS n FROM companies').fetchone()['n']
        if int(count) <= 1:
            raise ValueError('至少要保留一家公司，無法刪除最後一家公司。')
        record_count = self.conn.execute('SELECT COUNT(*) AS n FROM records WHERE company_id = ?', (company_id,)).fetchone()['n']
        manufacturer_count = self.conn.execute('SELECT COUNT(*) AS n FROM production_manufacturers WHERE company_id = ?', (company_id,)).fetchone()['n']
        if int(record_count) or int(manufacturer_count):
            raise ValueError(f'「{company["name"]}」仍有資料，為避免誤刪，公司內還有 {int(record_count)} 筆主資料與 {int(manufacturer_count)} 個廠商，請先清空後再刪除。')
        position_row = self.conn.execute('SELECT position FROM companies WHERE id = ?', (company_id,)).fetchone()
        position = int(position_row['position'])
        self.conn.execute('DELETE FROM companies WHERE id = ?', (company_id,))
        self.conn.execute('UPDATE companies SET position = position - 1 WHERE position > ?', (position,))
        self.conn.commit()

    def create_record(self, values: dict[int, Any], company_id: int | None = None) -> int:
        company_id = int(company_id or self.default_company_id())
        now = datetime.now().isoformat(timespec='seconds')
        cursor = self.conn.execute(
            'INSERT INTO records(created_at, updated_at, company_id) VALUES (?, ?, ?)', (now, now, company_id)
        )
        record_id = int(cursor.lastrowid)
        for field_id, value in values.items():
            self.set_value(record_id, int(field_id), '' if value is None else str(value))
        self._sync_linked_term(record_id)
        self.conn.commit()
        return record_id

    def update_record(self, record_id: int, values: dict[int, Any]) -> None:
        for field_id, value in values.items():
            self.set_value(record_id, int(field_id), '' if value is None else str(value))
        self._sync_linked_term(record_id)
        self.conn.execute(
            'UPDATE records SET updated_at = ? WHERE id = ?',
            (datetime.now().isoformat(timespec='seconds'), record_id),
        )
        self.conn.commit()

    def delete_record(self, record_id: int) -> None:
        count = self.production_record_count(record_id)
        if count:
            raise ValueError(f'此資料仍有 {count} 筆生產記錄，請先刪除或處理生產記錄後再刪除主資料。')
        self.conn.execute('DELETE FROM records WHERE id = ?', (record_id,))
        self.conn.commit()

    def set_value(self, record_id: int, field_id: int, value: str) -> None:
        self.conn.execute(
            '''INSERT INTO record_values(record_id, field_id, value) VALUES (?, ?, ?)
               ON CONFLICT(record_id, field_id) DO UPDATE SET value = excluded.value''',
            (record_id, field_id, value),
        )

    def get_record_values(self, record_id: int) -> dict[int, str]:
        rows = self.conn.execute(
            'SELECT field_id, value FROM record_values WHERE record_id = ?', (record_id,)
        )
        return {int(row['field_id']): (row['value'] or '') for row in rows}

    def get_record_links(self, record_id: int) -> dict[int, str]:
        rows = self.conn.execute(
            'SELECT field_id, url FROM record_links WHERE record_id = ?', (record_id,)
        )
        return {int(row['field_id']): row['url'] for row in rows}

    def set_record_link(self, record_id: int, field_id: int, url: str) -> None:
        url = url.strip()
        if url:
            self.conn.execute(
                'INSERT OR REPLACE INTO record_links(record_id, field_id, url) VALUES (?, ?, ?)',
                (int(record_id), int(field_id), url),
            )
        else:
            self.conn.execute(
                'DELETE FROM record_links WHERE record_id = ? AND field_id = ?',
                (int(record_id), int(field_id)),
            )
        self.conn.commit()

    def all_records(self, company_id: int | None = None) -> list[dict[str, Any]]:
        company_id = int(company_id or self.default_company_id())
        records = self.conn.execute('SELECT id, created_at, updated_at FROM records WHERE company_id = ? ORDER BY id', (company_id,))
        result = []
        for record in records:
            result.append({
                'id': int(record['id']),
                'created_at': record['created_at'],
                'updated_at': record['updated_at'],
                'values': self.get_record_values(int(record['id'])),
            })
        return result

    @staticmethod
    def _production_mismatch_sql() -> str:
        order_expr = "REPLACE(TRIM(COALESCE(order_quantity, quantity, '')), ',', '')"
        stock_expr = "REPLACE(TRIM(COALESCE(stock_quantity, '')), ',', '')"
        return f"(({order_expr} <> {stock_expr}) AND NOT ({order_expr} = '' AND {stock_expr} = ''))"

    def production_records(self, record_id: int, limit: int | None = None, offset: int = 0, abnormal_only: bool = False) -> list[sqlite3.Row]:
        sql = 'SELECT * FROM production_records WHERE record_id = ?'
        params: list[Any] = [int(record_id)]
        if abnormal_only:
            sql += ' AND ' + self._production_mismatch_sql()
        sql += ' ORDER BY production_date DESC, id DESC'
        if limit is not None:
            sql += ' LIMIT ? OFFSET ?'
            params.extend([max(1, int(limit)), max(0, int(offset))])
        return list(self.conn.execute(sql, tuple(params)))

    def production_record(self, production_id: int) -> sqlite3.Row | None:
        return self.conn.execute('SELECT * FROM production_records WHERE id = ?', (int(production_id),)).fetchone()

    def production_record_count(self, record_id: int, abnormal_only: bool = False) -> int:
        sql = 'SELECT COUNT(*) FROM production_records WHERE record_id = ?'
        params: list[Any] = [int(record_id)]
        if abnormal_only:
            sql += ' AND ' + self._production_mismatch_sql()
        return int(self.conn.execute(sql, tuple(params)).fetchone()[0])

    def create_production_record(self, record_id: int, production_date: str = '', order_quantity: str = '', stock_quantity: str = '', manufacturer: str = '', remark: str = '', batch_no: str = '', external_url: str = '') -> int:
        if not self.conn.execute('SELECT 1 FROM records WHERE id = ?', (int(record_id),)).fetchone():
            raise ValueError('找不到對應的主資料。')
        now = datetime.now().isoformat(timespec='seconds')
        cursor = self.conn.execute(
            '''INSERT INTO production_records(record_id, production_date, batch_no, quantity, order_quantity, stock_quantity, manufacturer, remark, external_url, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (int(record_id), production_date.strip(), batch_no.strip(), order_quantity.strip(), order_quantity.strip(), stock_quantity.strip(), manufacturer.strip(), remark.strip(), external_url.strip(), now, now),
        )
        self.conn.commit()
        return int(cursor.lastrowid)

    def update_production_record(self, production_id: int, production_date: str, order_quantity: str, stock_quantity: str, manufacturer: str, remark: str, batch_no: str = '', external_url: str = '') -> None:
        self.conn.execute(
            '''UPDATE production_records
               SET production_date = ?, batch_no = ?, quantity = ?, order_quantity = ?, stock_quantity = ?, manufacturer = ?, remark = ?, external_url = ?, updated_at = ?
               WHERE id = ?''',
            (production_date.strip(), batch_no.strip(), order_quantity.strip(), order_quantity.strip(), stock_quantity.strip(), manufacturer.strip(), remark.strip(), external_url.strip(), datetime.now().isoformat(timespec='seconds'), int(production_id)),
        )
        self.conn.commit()

    def delete_production_record(self, production_id: int) -> None:
        self.conn.execute('DELETE FROM production_records WHERE id = ?', (int(production_id),))
        self.conn.commit()

    def production_manufacturers(self, company_id: int | None = None) -> list[sqlite3.Row]:
        company_id = int(company_id or self.default_company_id())
        return list(self.conn.execute('SELECT * FROM production_manufacturers WHERE company_id = ? ORDER BY position, id', (company_id,)))

    def ensure_production_manufacturer(self, name: str, company_id: int | None = None) -> None:
        name = name.strip()
        company_id = int(company_id or self.default_company_id())
        if not name:
            return
        exists = self.conn.execute('SELECT 1 FROM production_manufacturers WHERE name = ? AND company_id = ?', (name, company_id)).fetchone()
        if not exists:
            position = int(self.conn.execute('SELECT COALESCE(MAX(position), -1) + 1 FROM production_manufacturers WHERE company_id = ?', (company_id,)).fetchone()[0])
            self.conn.execute('INSERT INTO production_manufacturers(name, company_id, position) VALUES (?, ?, ?)', (name, company_id, position))
            self.conn.commit()

    def add_production_manufacturer(self, name: str, company_id: int | None = None) -> None:
        name = name.strip()
        company_id = int(company_id or self.default_company_id())
        if not name:
            raise ValueError('製作廠商名稱不可空白。')
        if self.conn.execute('SELECT 1 FROM production_manufacturers WHERE name = ? AND company_id = ?', (name, company_id)).fetchone():
            raise ValueError('這個製作廠商已經存在。')
        position = int(self.conn.execute('SELECT COALESCE(MAX(position), -1) + 1 FROM production_manufacturers WHERE company_id = ?', (company_id,)).fetchone()[0])
        self.conn.execute('INSERT INTO production_manufacturers(name, company_id, position) VALUES (?, ?, ?)', (name, company_id, position))
        self.conn.commit()

    def update_production_manufacturer(self, manufacturer_id: int, name: str) -> None:
        name = name.strip()
        if not name:
            raise ValueError('製作廠商名稱不可空白。')
        duplicate = self.conn.execute(
            'SELECT 1 FROM production_manufacturers WHERE name = ? AND id <> ? AND company_id = (SELECT company_id FROM production_manufacturers WHERE id = ?)', (name, int(manufacturer_id), int(manufacturer_id))
        ).fetchone()
        if duplicate:
            raise ValueError('這個製作廠商已經存在。')
        old_row = self.conn.execute('SELECT name, company_id FROM production_manufacturers WHERE id = ?', (int(manufacturer_id),)).fetchone()
        old_name = str(old_row['name']) if old_row else ''
        company_id = int(old_row['company_id']) if old_row else self.default_company_id()
        self.conn.execute('UPDATE production_manufacturers SET name = ? WHERE id = ?', (name, int(manufacturer_id)))
        if old_name and old_name != name:
            self.conn.execute(
                '''UPDATE production_records SET manufacturer = ?
                   WHERE manufacturer = ?
                     AND record_id IN (SELECT id FROM records WHERE company_id = ?)''',
                (name, old_name, company_id),
            )
        self.conn.commit()

    def delete_production_manufacturer(self, manufacturer_id: int) -> None:
        row = self.conn.execute('SELECT name FROM production_manufacturers WHERE id = ?', (int(manufacturer_id),)).fetchone()
        if not row:
            return
        name = str(row['name'])
        company_id = int(row['company_id'])
        used = self.conn.execute(
            '''SELECT COUNT(*) FROM production_records
               WHERE manufacturer = ?
                 AND record_id IN (SELECT id FROM records WHERE company_id = ?)''',
            (name, company_id),
        ).fetchone()[0]
        if int(used) > 0:
            raise ValueError(f'「{name}」目前有 {int(used)} 筆生產記錄正在使用，請先修改那些記錄後再刪除。')
        self.conn.execute('DELETE FROM production_manufacturers WHERE id = ?', (int(manufacturer_id),))
        self.conn.commit()

    def lookup_pairs(self) -> list[sqlite3.Row]:
        return list(self.conn.execute('SELECT * FROM lookup_pairs ORDER BY position, id'))

    def add_lookup(self, code: str, term: str) -> None:
        code, term = code.strip(), term.strip()
        if not code or not term:
            raise ValueError('代碼與中文名稱皆不可空白。')
        if self.conn.execute('SELECT 1 FROM lookup_pairs WHERE code = ?', (code,)).fetchone():
            raise ValueError('代碼已存在。')
        position = self.conn.execute('SELECT COALESCE(MAX(position), -1) + 1 FROM lookup_pairs').fetchone()[0]
        self.conn.execute('INSERT INTO lookup_pairs(code, term, position) VALUES (?, ?, ?)', (code, term, position))
        self.conn.commit()

    def update_lookup(self, lookup_id: int, code: str, term: str) -> None:
        code, term = code.strip(), term.strip()
        if not code or not term:
            raise ValueError('代碼與中文名稱皆不可空白。')
        duplicate = self.conn.execute(
            'SELECT 1 FROM lookup_pairs WHERE code = ? AND id <> ?', (code, lookup_id)
        ).fetchone()
        if duplicate:
            raise ValueError('代碼已存在。')
        self.conn.execute('UPDATE lookup_pairs SET code = ?, term = ? WHERE id = ?', (code, term, lookup_id))
        self.conn.commit()

    def delete_lookup(self, lookup_id: int) -> None:
        self.conn.execute('DELETE FROM lookup_pairs WHERE id = ?', (lookup_id,))
        self.conn.commit()

    def _sync_linked_term(self, record_id: int) -> None:
        code_field = self.conn.execute(
            'SELECT id FROM fields WHERE system_key = ?', ('formulation_code',)
        ).fetchone()
        term_field = self.conn.execute(
            'SELECT id FROM fields WHERE system_key = ?', ('formulation_term',)
        ).fetchone()
        if not code_field or not term_field:
            return
        code = self.conn.execute(
            'SELECT value FROM record_values WHERE record_id = ? AND field_id = ?',
            (record_id, code_field['id']),
        ).fetchone()
        code_value = code['value'] if code else ''
        pair = self.conn.execute('SELECT term FROM lookup_pairs WHERE code = ?', (code_value,)).fetchone()
        term = pair['term'] if pair else ''
        self.set_value(record_id, int(term_field['id']), term)

    def backup(self, destination: Path | None = None) -> Path:
        self.conn.commit()
        if destination is None:
            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            destination = self.backup_dir / f'app_{stamp}.db'
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)
        copy_file_portable(self.path, destination)
        return destination

    def restore(self, source: Path) -> None:
        self.conn.close()
        copy_file_portable(source, self.path)
        self.conn = sqlite3.connect(self.path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute('PRAGMA foreign_keys = ON')
        self._init_schema()

    def close(self) -> None:
        if self.conn:
            self.conn.close()


# -----------------------------
# Small dialogs
# -----------------------------
class LookupDialog(tk.Toplevel):
    def __init__(self, parent: tk.Misc, title: str, code: str = '', term: str = ''):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.result: tuple[str, str] | None = None
        self.code_var = tk.StringVar(value=code)
        self.term_var = tk.StringVar(value=term)
        self._body()
        self.transient(parent)
        self.grab_set()
        self.protocol('WM_DELETE_WINDOW', self.destroy)

    def _body(self) -> None:
        frame = ttk.Frame(self, padding=16)
        frame.grid(sticky='nsew')
        ttk.Label(frame, text='英文代碼').grid(row=0, column=0, sticky='w', padx=(0, 8), pady=6)
        ttk.Entry(frame, textvariable=self.code_var, width=28).grid(row=0, column=1, pady=6)
        ttk.Label(frame, text='中文名稱').grid(row=1, column=0, sticky='w', padx=(0, 8), pady=6)
        ttk.Entry(frame, textvariable=self.term_var, width=28).grid(row=1, column=1, pady=6)
        buttons = ttk.Frame(frame)
        buttons.grid(row=2, column=0, columnspan=2, pady=(12, 0), sticky='e')
        ttk.Button(buttons, text='取消', command=self.destroy).pack(side='right', padx=(8, 0))
        ttk.Button(buttons, text='儲存', command=self._save).pack(side='right')

    def _save(self) -> None:
        self.result = (self.code_var.get(), self.term_var.get())
        self.destroy()


class FieldDialog(tk.Toplevel):
    def __init__(self, parent: tk.Misc, db: Database, field: sqlite3.Row | None = None):
        super().__init__(parent)
        self.db = db
        self.field = field
        self.title('新增欄位' if field is None else '編輯欄位設定')
        self.resizable(False, False)
        self.result = False
        self.name_var = tk.StringVar(value='' if field is None else field['name'])
        self.kind_var = tk.StringVar(value='text' if field is None else field['kind'])
        options = [] if field is None else json.loads(field['options_json'] or '[]')
        self.options_var = tk.StringVar(value=', '.join(options))
        self._body()
        self.transient(parent)
        self.grab_set()

    def _body(self) -> None:
        frame = ttk.Frame(self, padding=16)
        frame.grid(sticky='nsew')
        ttk.Label(frame, text='欄位名稱').grid(row=0, column=0, sticky='w', padx=(0, 8), pady=6)
        name_entry = ttk.Entry(frame, textvariable=self.name_var, width=36)
        name_entry.grid(row=0, column=1, pady=6)
        self.name_entry = name_entry
        ttk.Label(frame, text='欄位類型').grid(row=1, column=0, sticky='w', padx=(0, 8), pady=6)
        kinds = [('text', '文字輸入'), ('dropdown', '下拉選單')]
        kind_box = ttk.Combobox(frame, textvariable=self.kind_var, values=[x[0] for x in kinds], state='readonly', width=33)
        kind_box.grid(row=1, column=1, pady=6)
        kind_box.bind('<<ComboboxSelected>>', lambda _event: self._toggle_options())
        if self.field is not None and self.field['builtin']:
            kind_box.configure(state='disabled')
        ttk.Label(frame, text='選項').grid(row=2, column=0, sticky='nw', padx=(0, 8), pady=6)
        self.options_entry = ttk.Entry(frame, textvariable=self.options_var, width=36)
        self.options_entry.grid(row=2, column=1, pady=6)
        ttk.Label(frame, text='下拉選單請用逗號分隔，例如：甲,乙,丙', foreground='#666666').grid(row=3, column=1, sticky='w')
        buttons = ttk.Frame(frame)
        buttons.grid(row=4, column=0, columnspan=2, pady=(14, 0), sticky='e')
        if self.field is not None and self.field['system_key'] == 'formulation':
            ttk.Button(buttons, text='設定劑型底色', command=self._colors).pack(side='left', padx=(0, 20))
        ttk.Button(buttons, text='取消', command=self.destroy).pack(side='right', padx=(8, 0))
        ttk.Button(buttons, text='儲存', command=self._save).pack(side='right')
        self._toggle_options()

    def _toggle_options(self) -> None:
        if self.kind_var.get() == 'text':
            self.options_entry.configure(state='disabled')
        else:
            self.options_entry.configure(state='normal')

    def _colors(self) -> None:
        if self.field is None:
            return
        colors = json.loads(self.field['color_map_json'] or '{}')
        options = json.loads(self.field['options_json'] or '[]')
        ColorDialog(self, self.db, self.field['id'], options, colors)

    def _save(self) -> None:
        name = self.name_var.get().strip()
        if not name:
            messagebox.showerror('欄位設定', '欄位名稱不可空白。', parent=self)
            return
        options = [item.strip() for item in self.options_var.get().split(',') if item.strip()]
        try:
            if self.field is None:
                self.db.add_field(name, self.kind_var.get(), options)
            else:
                if self.field['builtin']:
                    self.db.rename_field(self.field['id'], name)
                    if self.field['kind'] != 'text':
                        self.db.update_field_options(self.field['id'], options)
                else:
                    self.db.conn.execute(
                        'UPDATE fields SET name = ?, kind = ?, options_json = ? WHERE id = ?',
                        (name, self.kind_var.get(), json.dumps(options, ensure_ascii=False), self.field['id']),
                    )
                    self.db.conn.commit()
            self.result = True
            self.destroy()
        except ValueError as exc:
            messagebox.showerror('欄位設定', str(exc), parent=self)


class LinkDialog(tk.Toplevel):
    def __init__(self, parent: tk.Misc, current_url: str = ''):
        super().__init__(parent)
        self.title('編輯連結')
        self.resizable(False, False)
        self.result: str | None = None
        self.url_var = tk.StringVar(value=current_url)
        frame = ttk.Frame(self, padding=16)
        frame.grid(sticky='nsew')
        ttk.Label(frame, text='連結網址或檔案路徑').grid(row=0, column=0, sticky='w', padx=(0, 8), pady=6)
        entry = ttk.Entry(frame, textvariable=self.url_var, width=54)
        entry.grid(row=0, column=1, pady=6)
        entry.focus_set()
        buttons = ttk.Frame(frame)
        buttons.grid(row=1, column=0, columnspan=2, sticky='e', pady=(12, 0))
        ttk.Button(buttons, text='取消', command=self.destroy).pack(side='right', padx=(8, 0))
        ttk.Button(buttons, text='儲存連結', command=self._save).pack(side='right')
        self.bind('<Return>', lambda _event: self._save())
        self.bind('<Escape>', lambda _event: self.destroy())
        self.transient(parent)
        self.grab_set()

    def _save(self) -> None:
        self.result = self.url_var.get().strip()
        self.destroy()


class ColorDialog(tk.Toplevel):
    def __init__(self, parent: tk.Misc, db: Database, field_id: int, options: list[str], colors: dict[str, str]):
        super().__init__(parent)
        self.db = db
        self.field_id = field_id
        self.colors = dict(colors)
        self.title('設定劑型底色')
        self.resizable(False, False)
        frame = ttk.Frame(self, padding=16)
        frame.grid(sticky='nsew')
        self.entries: dict[str, tk.StringVar] = {}
        for row, option in enumerate(options):
            ttk.Label(frame, text=option, width=12).grid(row=row, column=0, sticky='w', pady=5)
            var = tk.StringVar(value=self.colors.get(option, '#FFFFFF'))
            self.entries[option] = var
            ttk.Entry(frame, textvariable=var, width=14).grid(row=row, column=1, padx=6, pady=5)
            ttk.Button(frame, text='選色', command=lambda opt=option, v=var: self._choose(opt, v)).grid(row=row, column=2, pady=5)
        buttons = ttk.Frame(frame)
        buttons.grid(row=len(options), column=0, columnspan=3, sticky='e', pady=(12, 0))
        ttk.Button(buttons, text='取消', command=self.destroy).pack(side='right', padx=(8, 0))
        ttk.Button(buttons, text='儲存', command=self._save).pack(side='right')
        self.transient(parent)
        self.grab_set()

    def _choose(self, option: str, var: tk.StringVar) -> None:
        result = colorchooser.askcolor(color=var.get(), title=f'選擇「{option}」底色', parent=self)
        if result and result[1]:
            var.set(result[1].upper())

    def _save(self) -> None:
        colors = {option: var.get().strip() for option, var in self.entries.items()}
        for option, color in colors.items():
            if not (color.startswith('#') and len(color) == 7):
                messagebox.showerror('底色設定', f'「{option}」的顏色格式應為 #RRGGBB。', parent=self)
                return
        self.db.update_field_colors(self.field_id, colors)
        self.destroy()


class DisplaySettingsDialog(tk.Toplevel):
    def __init__(self, parent: tk.Misc, db: Database, settings: dict[str, Any]):
        super().__init__(parent)
        self.db = db
        self.settings = dict(settings)
        self.title('顯示設定')
        self.geometry('460x370')
        self.minsize(420, 340)
        self.resizable(False, False)
        self.result = False
        self.header_font_size_var = tk.IntVar(value=int(settings.get('header_font_size', settings.get('font_size', 10))))
        self.data_font_size_var = tk.IntVar(value=int(settings.get('data_font_size', settings.get('font_size', 10))))
        self.header_row_height_var = tk.IntVar(value=int(settings.get('header_row_height', 42)))
        self.data_row_height_var = tk.IntVar(value=int(settings.get('data_row_height', 28)))
        self.link_marker_var = tk.StringVar(value=str(settings.get('link_marker', LINK_MARKER)))
        self.show_link_marker_var = tk.BooleanVar(value=bool(settings.get('show_link_marker', True)))
        self._body()
        self.transient(parent)
        self.grab_set()

    def _body(self) -> None:
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill='both', expand=True)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text='標題字體大小').grid(row=0, column=0, sticky='w', pady=7)
        ttk.Spinbox(frame, from_=8, to=24, textvariable=self.header_font_size_var, width=12).grid(row=0, column=1, sticky='w', pady=7)
        ttk.Label(frame, text='資料字體大小').grid(row=1, column=0, sticky='w', pady=7)
        ttk.Spinbox(frame, from_=8, to=24, textvariable=self.data_font_size_var, width=12).grid(row=1, column=1, sticky='w', pady=7)
        ttk.Label(frame, text='標題列高').grid(row=2, column=0, sticky='w', pady=7)
        ttk.Spinbox(frame, from_=24, to=100, textvariable=self.header_row_height_var, width=12).grid(row=2, column=1, sticky='w', pady=7)
        ttk.Label(frame, text='資料列高').grid(row=3, column=0, sticky='w', pady=7)
        ttk.Spinbox(frame, from_=20, to=80, textvariable=self.data_row_height_var, width=12).grid(row=3, column=1, sticky='w', pady=7)
        ttk.Label(frame, text='連結圖示文字').grid(row=4, column=0, sticky='w', pady=7)
        ttk.Entry(frame, textvariable=self.link_marker_var, width=12).grid(row=4, column=1, sticky='w', pady=7)
        ttk.Checkbutton(frame, text='顯示資料表連結圖示', variable=self.show_link_marker_var).grid(row=5, column=0, columnspan=2, sticky='w', pady=4)
        ttk.Label(frame, text='可輸入 1 個符號或短文字；按「儲存設定」後立即套用並保存。', style='Hint.TLabel').grid(row=6, column=0, columnspan=2, sticky='w', pady=(4, 10))

        buttons = ttk.Frame(frame)
        buttons.grid(row=7, column=0, columnspan=2, sticky='e', pady=(4, 0))
        ttk.Button(buttons, text='取消', command=self.destroy).pack(side='right', padx=(8, 0))
        ttk.Button(buttons, text='儲存設定', command=self._save).pack(side='right')

    def _save(self) -> None:
        try:
            header_font_size = max(8, min(24, int(self.header_font_size_var.get())))
            data_font_size = max(8, min(24, int(self.data_font_size_var.get())))
            header_row_height = max(24, min(100, int(self.header_row_height_var.get())))
            data_row_height = max(20, min(80, int(self.data_row_height_var.get())))
            link_marker = self.link_marker_var.get().strip() or LINK_MARKER
            if len(link_marker) > 4:
                raise ValueError
        except (TypeError, ValueError, tk.TclError):
            messagebox.showerror('顯示設定', '請輸入有效的數字。', parent=self)
            return
        self.settings.update({
            'header_font_size': header_font_size,
            'data_font_size': data_font_size,
            'header_row_height': header_row_height,
            'data_row_height': data_row_height,
            'link_marker': link_marker,
            'show_link_marker': bool(self.show_link_marker_var.get()),
        })
        self.db.save_display_settings(self.settings)
        self.result = True
        self.destroy()


# -----------------------------
# Main application
# -----------------------------
class OfflineDatabaseApp(tk.Tk):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.display_settings = self.db.get_display_settings()
        self.title(APP_TITLE)
        self.wm_title(APP_TITLE)
        try:
            self.iconbitmap(default='')
        except tk.TclError:
            pass
        # 以透明 1x1 icon 隱藏 Windows 標題列左側的預設 Tk icon。
        self._blank_window_icon = tk.PhotoImage(width=1, height=1)
        self.iconphoto(True, self._blank_window_icon)
        self.geometry('1440x820')
        self.minsize(1050, 650)
        self.protocol('WM_DELETE_WINDOW', self._on_close)
        self.current_company_id = self.db.default_company_id()
        self.company_notebook: ttk.Notebook | None = None
        self.search_var = tk.StringVar()
        self.status_var = tk.StringVar(value='就緒')
        self.tree: ttk.Treeview
        self._sort_field_id: int | None = None
        self._sort_reverse = False
        self._build_style()
        self._build_ui()
        self.refresh_all()

    def _build_style(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use('vista')
        except tk.TclError:
            pass
        header_font_size = int(self.display_settings.get('header_font_size', self.display_settings.get('font_size', 10)))
        data_font_size = int(self.display_settings.get('data_font_size', self.display_settings.get('font_size', 10)))
        header_row_height = int(self.display_settings.get('header_row_height', 42))
        data_row_height = int(self.display_settings.get('data_row_height', 28))
        self._tree_font = ('Microsoft JhengHei UI', data_font_size)
        self._tree_heading_font = ('Microsoft JhengHei UI', header_font_size, 'bold')
        style.configure('Title.TLabel', font=('Microsoft JhengHei UI', FIXED_UI_FONT_SIZE + 8, 'bold'))
        style.configure('Hint.TLabel', foreground='#666666', font=('Microsoft JhengHei UI', FIXED_UI_FONT_SIZE))
        style.configure('Treeview', rowheight=data_row_height, font=self._tree_font, borderwidth=0, relief='flat')
        style.map('Treeview', background=[('selected', '#FFFFFF')], foreground=[('selected', '#000000')])
        style.layout('Treeview', [('Treeview.treearea', {'sticky': 'nswe'})])
        style.layout('FieldTreeview', style.layout('Treeview'))
        style.configure('FieldTreeview', rowheight=data_row_height, font=self._tree_font, borderwidth=0, relief='flat')
        style.map('FieldTreeview', background=[('selected', '#000000')], foreground=[('selected', '#FFFFFF')])
        style.layout('FieldTreeview', [('Treeview.treearea', {'sticky': 'nswe'})])
        style.configure('Treeview.Heading', font=self._tree_heading_font, padding=(6, max(4, (header_row_height - 20) // 2)))

    def _build_ui(self) -> None:
        header = ttk.Frame(self, padding=(16, 12, 16, 6))
        header.pack(fill='x')
        ttk.Label(header, text=APP_TITLE, style='Title.TLabel', anchor='w').pack(side='left', fill='x', expand=False)
        ttk.Label(header, text='SQLite／完全離線／單一使用者', style='Hint.TLabel').pack(side='left', padx=18)
        ttk.Label(header, textvariable=self.status_var, style='Hint.TLabel').pack(side='right')

        notebook = ttk.Notebook(self)
        notebook.pack(fill='both', expand=True, padx=12, pady=(0, 12))
        self.data_tab = ttk.Frame(notebook)
        self.fields_tab = ttk.Frame(notebook)
        self.lookup_tab = ttk.Frame(notebook)
        self.backup_tab = ttk.Frame(notebook)
        notebook.add(self.data_tab, text='資料管理')
        notebook.add(self.fields_tab, text='欄位管理')
        notebook.add(self.lookup_tab, text='代碼／名稱對照')
        notebook.add(self.backup_tab, text='備份與還原')
        self._build_data_tab()
        self._build_fields_tab()
        self._build_lookup_tab()
        self._build_backup_tab()

    def _build_data_tab(self) -> None:
        company_bar = ttk.Frame(self.data_tab, padding=(10, 10, 10, 4))
        company_bar.pack(fill='x')
        ttk.Label(company_bar, text='公司').pack(side='left', padx=(0, 8))
        self.company_notebook = ttk.Notebook(company_bar, height=34)
        self.company_notebook.pack(side='left', fill='x', expand=False)
        for company in self.db.companies():
            frame = ttk.Frame(self.company_notebook)
            self.company_notebook.add(frame, text=str(company['name']))
        self.company_notebook.bind('<<NotebookTabChanged>>', self._on_company_tab_changed)
        companies = self.db.companies()
        for idx, company in enumerate(companies):
            if int(company['id']) == self.current_company_id:
                self.company_notebook.select(idx)
                break
        ttk.Button(company_bar, text='公司管理', command=self.manage_companies).pack(side='left', padx=10)
        ttk.Label(company_bar, text='目前公司：', style='Hint.TLabel').pack(side='left', padx=(8, 0))
        self.company_label_var = tk.StringVar()
        ttk.Label(company_bar, textvariable=self.company_label_var, font=('Microsoft JhengHei UI', 10, 'bold')).pack(side='left', padx=(4, 0))

        toolbar = ttk.Frame(self.data_tab, padding=(10, 6, 10, 6))
        toolbar.pack(fill='x')
        ttk.Label(toolbar, text='搜尋').pack(side='left')
        search = ttk.Entry(toolbar, textvariable=self.search_var, width=34)
        search.pack(side='left', padx=(8, 5))
        search.bind('<Return>', lambda _event: self.refresh_data())
        ttk.Button(toolbar, text='搜尋', command=self.refresh_data).pack(side='left', padx=(0, 18))
        ttk.Button(toolbar, text='新增資料', command=lambda: self.open_record_editor(None)).pack(side='left', padx=3)
        ttk.Button(toolbar, text='編輯資料', command=self.edit_selected_record).pack(side='left', padx=3)
        ttk.Button(toolbar, text='刪除資料', command=self.delete_selected_record).pack(side='left', padx=3)
        ttk.Button(toolbar, text='從 Excel 匯入／更新', command=self.import_excel).pack(side='left', padx=18)
        ttk.Button(toolbar, text='重新整理', command=self.refresh_data).pack(side='left', padx=3)
        ttk.Button(toolbar, text='顯示設定', command=self.open_display_settings).pack(side='left', padx=(18, 3))
        ttk.Label(toolbar, text='點擊欄位標題可排序', style='Hint.TLabel').pack(side='left', padx=10)

        table_frame = ttk.Frame(self.data_tab, padding=(10, 0, 10, 10))
        table_frame.pack(fill='both', expand=True)
        self.tree = ttk.Treeview(table_frame, show='headings', selectmode='browse')
        yscroll = ttk.Scrollbar(table_frame, orient='vertical', command=self.tree.yview)
        xscroll = ttk.Scrollbar(table_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        yscroll.grid(row=0, column=1, sticky='ns')
        xscroll.grid(row=1, column=0, sticky='ew')
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self.tree.bind('<Double-1>', self._on_tree_double_click)
        self.tree.bind('<ButtonRelease-1>', self._on_tree_button_release, add='+')
        self.tree.bind('<ButtonRelease-1>', self._on_tree_link_click, add='+')
        self.tree.bind('<ButtonRelease-1>', self._on_tree_production_click, add='+')
        self._on_company_tab_changed()

    def _on_company_tab_changed(self, _event: Any = None) -> None:
        if not self.company_notebook:
            return
        tabs = self.company_notebook.tabs()
        selected = self.company_notebook.select()
        if selected in tabs:
            index = tabs.index(selected)
            companies = self.db.companies()
            if index < len(companies):
                self.current_company_id = int(companies[index]['id'])
                if hasattr(self, 'company_label_var'):
                    self.company_label_var.set(str(companies[index]['name']))
                self._sort_field_id = None
                self._sort_reverse = False
                self.refresh_data()
                self.status_var.set(f'目前公司：{companies[index]["name"]}')

    def manage_companies(self) -> None:
        dialog = tk.Toplevel(self)
        dialog.title('公司管理')
        dialog.geometry('520x390')
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        body = ttk.Frame(dialog, padding=16)
        body.pack(fill='both', expand=True)
        ttk.Label(body, text='公司管理', style='Title.TLabel').pack(anchor='w')
        ttk.Label(body, text='可新增、修改或刪除公司。刪除公司前，必須先清空該公司的資料。', style='Hint.TLabel').pack(anchor='w', pady=(4, 12))

        list_frame = ttk.Frame(body)
        list_frame.pack(fill='both', expand=True)
        company_tree = ttk.Treeview(list_frame, columns=('name', 'count'), show='headings', height=9, selectmode='browse')
        company_tree.heading('name', text='公司名稱')
        company_tree.heading('count', text='主資料筆數')
        company_tree.column('name', width=300, anchor='w')
        company_tree.column('count', width=120, anchor='center')
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=company_tree.yview)
        company_tree.configure(yscrollcommand=scrollbar.set)
        company_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        def reload_list(select_id: int | None = None):
            for item in company_tree.get_children():
                company_tree.delete(item)
            for company in self.db.companies():
                cid = int(company['id'])
                count = self.db.conn.execute('SELECT COUNT(*) AS n FROM records WHERE company_id = ?', (cid,)).fetchone()['n']
                company_tree.insert('', 'end', iid=str(cid), values=(str(company['name']), int(count)))
            if select_id is not None and company_tree.exists(str(select_id)):
                company_tree.selection_set(str(select_id))
                company_tree.focus(str(select_id))

        def selected_company():
            selected = company_tree.selection()
            if not selected:
                messagebox.showinfo('公司管理', '請先選擇一家公司。', parent=dialog)
                return None
            cid = int(selected[0])
            row = self.db.conn.execute('SELECT id, name FROM companies WHERE id = ?', (cid,)).fetchone()
            return row

        def add_company():
            name = simpledialog.askstring('新增公司', '請輸入公司名稱：', parent=dialog)
            if name is None:
                return
            try:
                cid = self.db.add_company(name)
                self.current_company_id = cid
                self._rebuild_company_tabs()
                reload_list(cid)
            except ValueError as exc:
                messagebox.showerror('新增公司', str(exc), parent=dialog)

        def rename_selected():
            row = selected_company()
            if not row:
                return
            name = simpledialog.askstring('修改公司名稱', '請輸入新的公司名稱：', initialvalue=str(row['name']), parent=dialog)
            if name is None:
                return
            try:
                self.db.rename_company(int(row['id']), name)
                self._rebuild_company_tabs()
                reload_list(int(row['id']))
            except ValueError as exc:
                messagebox.showerror('修改公司名稱', str(exc), parent=dialog)

        def delete_selected():
            row = selected_company()
            if not row:
                return
            cid = int(row['id'])
            name = str(row['name'])
            if not messagebox.askyesno('刪除公司', f'確定要刪除「{name}」嗎？\n\n只有完全沒有資料的公司才能刪除。', parent=dialog):
                return
            try:
                self.db.delete_company(cid)
                if self.current_company_id == cid:
                    self.current_company_id = self.db.default_company_id()
                self._rebuild_company_tabs()
                reload_list(self.current_company_id)
            except ValueError as exc:
                messagebox.showerror('刪除公司', str(exc), parent=dialog)

        buttons = ttk.Frame(body)
        buttons.pack(fill='x', pady=(12, 0))
        ttk.Button(buttons, text='新增公司', command=add_company).pack(side='left')
        ttk.Button(buttons, text='修改名稱', command=rename_selected).pack(side='left', padx=6)
        ttk.Button(buttons, text='刪除公司', command=delete_selected).pack(side='left')
        ttk.Button(buttons, text='關閉', command=dialog.destroy).pack(side='right')

        reload_list(self.current_company_id)
        dialog.wait_window()

    def edit_company_names(self) -> None:
        # 舊方法名稱保留，避免其他舊程式碼呼叫時失效。
        self.manage_companies()

    def _rebuild_company_tabs(self) -> None:
        if not self.company_notebook:
            return
        current_id = self.current_company_id
        for tab in self.company_notebook.tabs():
            frame = self.company_notebook.nametowidget(tab)
            self.company_notebook.forget(tab)
            frame.destroy()
        companies = self.db.companies()
        if not any(int(company['id']) == current_id for company in companies):
            current_id = self.db.default_company_id()
            self.current_company_id = current_id
        for company in companies:
            frame = ttk.Frame(self.company_notebook)
            self.company_notebook.add(frame, text=str(company['name']))
        companies = self.db.companies()
        for idx, company in enumerate(companies):
            if int(company['id']) == current_id:
                self.company_notebook.select(idx)
                break
        self._on_company_tab_changed()

    def _build_fields_tab(self) -> None:
        toolbar = ttk.Frame(self.fields_tab, padding=10)
        toolbar.pack(fill='x')
        ttk.Button(toolbar, text='新增欄位', command=self.add_field).pack(side='left')
        ttk.Button(toolbar, text='編輯設定', command=self.edit_selected_field).pack(side='left', padx=6)
        ttk.Button(toolbar, text='設定劑型底色', command=self.edit_category_colors).pack(side='left', padx=6)
        ttk.Button(toolbar, text='上移', command=lambda: self.move_selected_field(-1)).pack(side='left', padx=(18, 3))
        ttk.Button(toolbar, text='下移', command=lambda: self.move_selected_field(1)).pack(side='left', padx=3)
        ttk.Label(toolbar, text='可拖曳欄位列，或使用上移／下移調整順序。', style='Hint.TLabel').pack(side='left', padx=18)
        frame = ttk.Frame(self.fields_tab, padding=(10, 0, 10, 10))
        frame.pack(fill='both', expand=True)
        self.fields_tree = ttk.Treeview(frame, columns=('position', 'name', 'kind', 'options'), show='headings', selectmode='browse', style='FieldTreeview')
        for key, text, width in [('position', '順序', 70), ('name', '欄位名稱', 270), ('kind', '類型', 120), ('options', '下拉選項', 600)]:
            self.fields_tree.heading(key, text=text)
            self.fields_tree.column(key, width=width, anchor='w')
        scroll = ttk.Scrollbar(frame, orient='vertical', command=self.fields_tree.yview)
        self.fields_tree.configure(yscrollcommand=scroll.set)
        self.fields_tree.grid(row=0, column=0, sticky='nsew')
        scroll.grid(row=0, column=1, sticky='ns')
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        self.fields_tree.bind('<Double-1>', lambda _event: self.edit_selected_field())
        self.fields_tree.bind('<ButtonPress-1>', self._on_field_press, add='+')
        self.fields_tree.bind('<B1-Motion>', self._on_field_drag, add='+')
        self.fields_tree.bind('<ButtonRelease-1>', self._on_field_release, add='+')

    def _on_field_press(self, event: Any) -> None:
        iid = self.fields_tree.identify_row(event.y)
        self._field_drag_iid = iid or None
        if iid:
            self.fields_tree.selection_set(iid)

    def _on_field_drag(self, event: Any) -> None:
        source_iid = getattr(self, '_field_drag_iid', None)
        if not source_iid:
            return
        target_iid = self.fields_tree.identify_row(event.y)
        if not target_iid or target_iid == source_iid:
            return
        children = list(self.fields_tree.get_children())
        remaining = [item for item in children if item != source_iid]
        target_index = remaining.index(target_iid)
        target_bbox = self.fields_tree.bbox(target_iid)
        if target_bbox and event.y > target_bbox[1] + target_bbox[3] // 2:
            target_index += 1
        self.fields_tree.move(source_iid, '', target_index)
        self.fields_tree.selection_set(source_iid)

    def _on_field_release(self, _event: Any) -> None:
        source_iid = getattr(self, '_field_drag_iid', None)
        self._field_drag_iid = None
        if not source_iid:
            return
        ordered_ids = [int(iid) for iid in self.fields_tree.get_children()]
        self.db.reorder_fields(ordered_ids)
        self.refresh_fields()
        self.refresh_data()
        if self.fields_tree.exists(source_iid):
            self.fields_tree.selection_set(source_iid)

    def move_selected_field(self, direction: int) -> None:
        iid = self._selected_id(self.fields_tree)
        if iid is None:
            messagebox.showwarning('欄位管理', '請先選取要移動的欄位。', parent=self)
            return
        ordered_ids = [int(item) for item in self.fields_tree.get_children()]
        current_index = ordered_ids.index(iid)
        target_index = current_index + (1 if direction > 0 else -1)
        if target_index < 0 or target_index >= len(ordered_ids):
            return
        ordered_ids[current_index], ordered_ids[target_index] = ordered_ids[target_index], ordered_ids[current_index]
        self.db.reorder_fields(ordered_ids)
        self.refresh_fields()
        self.refresh_data()
        self.fields_tree.selection_set(str(iid))

    def _build_lookup_tab(self) -> None:
        toolbar = ttk.Frame(self.lookup_tab, padding=10)
        toolbar.pack(fill='x')
        ttk.Button(toolbar, text='新增對照', command=self.add_lookup).pack(side='left')
        ttk.Button(toolbar, text='編輯對照', command=self.edit_selected_lookup).pack(side='left', padx=6)
        ttk.Button(toolbar, text='刪除對照', command=self.delete_selected_lookup).pack(side='left', padx=6)
        ttk.Label(toolbar, text='G 欄選擇英文代碼後，F 欄會自動顯示對應中文名稱。', style='Hint.TLabel').pack(side='left', padx=18)
        frame = ttk.Frame(self.lookup_tab, padding=(10, 0, 10, 10))
        frame.pack(fill='both', expand=True)
        self.lookup_tree = ttk.Treeview(frame, columns=('code', 'term'), show='headings', selectmode='browse', style='FieldTreeview')
        self.lookup_tree.heading('code', text='英文代碼')
        self.lookup_tree.heading('term', text='中文名稱')
        self.lookup_tree.column('code', width=240, anchor='w')
        self.lookup_tree.column('term', width=360, anchor='w')
        scroll = ttk.Scrollbar(frame, orient='vertical', command=self.lookup_tree.yview)
        self.lookup_tree.configure(yscrollcommand=scroll.set)
        self.lookup_tree.grid(row=0, column=0, sticky='nsew')
        scroll.grid(row=0, column=1, sticky='ns')
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        self.lookup_tree.bind('<Double-1>', lambda _event: self.edit_selected_lookup())

    def _build_backup_tab(self) -> None:
        frame = ttk.Frame(self.backup_tab, padding=24)
        frame.pack(fill='both', expand=True, anchor='nw')
        ttk.Label(frame, text='資料庫檔案位置', font=('Microsoft JhengHei UI', 11, 'bold')).pack(anchor='w')
        ttk.Label(frame, text=str(self.db.path), style='Hint.TLabel').pack(anchor='w', pady=(6, 20))
        ttk.Label(frame, text='備份與還原會直接處理整個 SQLite 資料庫，建議定期備份 data\app.db。', style='Hint.TLabel').pack(anchor='w', pady=(0, 14))
        ttk.Button(frame, text='立即備份到程式 backup 資料夾', command=self.backup_now).pack(anchor='w', pady=5)
        ttk.Button(frame, text='另存備份檔', command=self.backup_as).pack(anchor='w', pady=5)
        ttk.Button(frame, text='從備份檔還原', command=self.restore_backup).pack(anchor='w', pady=5)

    def refresh_all(self) -> None:
        self.refresh_data()
        self.refresh_fields()
        self.refresh_lookup()

    def _on_tree_button_release(self, _event: Any) -> None:
        # 讓 Treeview 先完成拖曳寬度更新，再讀取實際欄寬並保存。
        self.after_idle(self._save_current_column_widths)

    def _link_at_event(self, event: Any) -> str | None:
        item = self.tree.identify_row(event.y)
        if not item:
            return None
        column = self.tree.identify_column(event.x)
        if not column.startswith('#'):
            return None
        try:
            column_index = int(column[1:]) - 1
            column_id = self.tree['columns'][column_index]
        except (ValueError, IndexError, tk.TclError):
            return None
        if not str(column_id).startswith('link'):
            return None
        try:
            field_id = int(str(column_id)[4:])
            record_id = int(item)
        except ValueError:
            return None
        return self.db.get_record_links(record_id).get(field_id)

    def _on_tree_link_click(self, event: Any) -> None:
        url = self._link_at_event(event)
        if url:
            self.open_link(url)

    def _production_record_at_event(self, event: Any) -> int | None:
        item = self.tree.identify_row(event.y)
        if not item:
            return None
        column = self.tree.identify_column(event.x)
        if not column.startswith('#'):
            return None
        try:
            column_index = int(column[1:]) - 1
            column_id = self.tree['columns'][column_index]
            if not str(column_id).startswith('production'):
                return None
            return int(item)
        except (ValueError, IndexError, tk.TclError):
            return None

    def _on_tree_production_click(self, event: Any) -> None:
        record_id = self._production_record_at_event(event)
        if record_id is not None:
            self.open_production_records(record_id)

    def _on_tree_double_click(self, event: Any) -> str | None:
        if self._production_record_at_event(event) is not None:
            return 'break'
        url = self._link_at_event(event)
        if url:
            self.open_link(url)
            return 'break'
        self.edit_selected_record()
        return None

    def _save_current_column_widths(self) -> None:
        if not hasattr(self, 'tree') or not self.tree.winfo_exists():
            return
        fields = self.db.list_fields()
        widths = dict(self.display_settings.get('column_widths', {}))
        link_widths = dict(self.display_settings.get('link_column_widths', {}))
        production_widths = dict(self.display_settings.get('production_column_widths', {}))
        changed = False
        for field in fields:
            field_id = int(field['id'])
            column_id = f'c{field_id}'
            try:
                width = int(self.tree.column(column_id, 'width'))
            except (tk.TclError, TypeError, ValueError):
                width = 0
            if width > 0 and widths.get(str(field_id)) != width:
                widths[str(field_id)] = width
                changed = True
            link_column_id = f'link{field_id}'
            try:
                link_width = int(self.tree.column(link_column_id, 'width'))
            except (tk.TclError, TypeError, ValueError):
                link_width = 0
            if link_width > 0 and link_widths.get(str(field_id)) != link_width:
                link_widths[str(field_id)] = link_width
                changed = True
            production_column_id = f'production{field_id}'
            try:
                production_width = int(self.tree.column(production_column_id, 'width'))
            except (tk.TclError, TypeError, ValueError):
                production_width = 0
            if production_width > 0 and production_widths.get(str(field_id)) != production_width:
                production_widths[str(field_id)] = production_width
                changed = True
        if changed:
            self.display_settings['column_widths'] = widths
            self.display_settings['link_column_widths'] = link_widths
            self.display_settings['production_column_widths'] = production_widths
            self.db.save_display_settings(self.display_settings)

    def refresh_data(self) -> None:
        fields = self.db.list_fields()
        records = self.db.all_records(self.current_company_id)
        record_links = {int(record['id']): self.db.get_record_links(int(record['id'])) for record in records}
        show_link_marker = bool(self.display_settings.get('show_link_marker', True))
        link_marker = str(self.display_settings.get('link_marker', LINK_MARKER)) or LINK_MARKER
        link_field_ids = {
            int(field['id'])
            for field in fields
            if show_link_marker and any(int(field['id']) in links for links in record_links.values())
        }
        query = self.search_var.get().strip().lower()
        self.tree.delete(*self.tree.get_children())
        columns: list[str] = []
        for field in fields:
            field_id = int(field['id'])
            columns.append(f'c{field_id}')
            if field_id in link_field_ids:
                columns.append(f'link{field_id}')
            if field['system_key'] == 'brand_name':
                columns.append(f'production{field_id}')
        self.tree['columns'] = columns
        column_widths = self.display_settings.get('column_widths', {})
        link_column_widths = self.display_settings.get('link_column_widths', {})
        production_widths = self.display_settings.get('production_column_widths', {})
        fallback_width = DEFAULT_COLUMN_WIDTH
        for field in fields:
            field_id = int(field['id'])
            column_id = f'c{field_id}'
            arrow = ''
            if self._sort_field_id == field_id:
                arrow = ' ▼' if self._sort_reverse else ' ▲'
            self.tree.heading(
                column_id,
                text=field['name'] + arrow,
                anchor='center',
                command=lambda selected_id=field_id: self.sort_by_field(selected_id),
            )
            width = int(column_widths.get(str(field_id), fallback_width))
            width = max(60, min(600, width))
            self.tree.column(column_id, width=width, anchor='center', stretch=False)
            if field_id in link_field_ids:
                link_column_id = f'link{field_id}'
                self.tree.heading(link_column_id, text='', anchor='center')
                link_width = int(link_column_widths.get(str(field_id), DEFAULT_LINK_COLUMN_WIDTH))
                link_width = max(18, min(100, link_width))
                self.tree.column(link_column_id, width=link_width, minwidth=18, anchor='center', stretch=False)
            if field['system_key'] == 'brand_name':
                production_column_id = f'production{field_id}'
                self.tree.heading(production_column_id, text='', anchor='center')
                production_width = int(production_widths.get(str(field_id), DEFAULT_PRODUCTION_COLUMN_WIDTH))
                production_width = max(28, min(100, production_width))
                self.tree.column(production_column_id, width=production_width, minwidth=28, anchor='center', stretch=False)
        if self._sort_field_id is not None:
            sort_field_id = self._sort_field_id
            records_nonempty = []
            records_empty = []
            for record in records:
                value = str(record['values'].get(sort_field_id, '') or '').strip()
                (records_empty if not value else records_nonempty).append(record)
            records_nonempty.sort(key=lambda record: self._sort_key(record['values'].get(sort_field_id, '')), reverse=self._sort_reverse)
            records = records_nonempty + records_empty
        category_field = next((f for f in fields if f['system_key'] == 'formulation'), None)
        category_colors = json.loads(category_field['color_map_json'] or '{}') if category_field else {}
        color_tags: dict[str, str] = {}
        shown = 0
        for record in records:
            values: list[Any] = []
            for field in fields:
                field_id = int(field['id'])
                values.append(record['values'].get(field_id, ''))
                if field['system_key'] == 'brand_name':
                    values.append(PRODUCTION_MARKER)
                if field_id in link_field_ids:
                    values.append(link_marker if record_links.get(int(record['id']), {}).get(field_id) else '')
            searchable_values = [record['values'].get(int(field['id']), '') for field in fields]
            if query and query not in ' '.join(str(value).lower() for value in searchable_values):
                continue
            tags: tuple[str, ...] = ()
            category_value = ''
            if category_field:
                category_value = record['values'].get(int(category_field['id']), '')
            if category_value in category_colors:
                tag = f'category_{category_value}'
                if tag not in color_tags:
                    color_tags[tag] = category_colors[category_value]
                    self.tree.tag_configure(tag, background=category_colors[category_value])
                tags = (tag,)
            self.tree.insert('', 'end', iid=str(record['id']), values=values, tags=tags)
            shown += 1
        self.status_var.set(f'資料 {shown} 筆／欄位 {len(fields)} 個')

    @staticmethod
    def _sort_key(value: Any) -> tuple[Any, ...]:
        text = str(value or '').strip()
        try:
            numeric = float(text.rstrip('%'))
            return (0, 0, numeric)
        except ValueError:
            return (0, 1, text.casefold())

    def sort_by_field(self, field_id: int) -> None:
        if self._sort_field_id == field_id:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_field_id = field_id
            self._sort_reverse = False
        self.refresh_data()

    def open_display_settings(self) -> None:
        dialog = DisplaySettingsDialog(self, self.db, self.display_settings)
        self.wait_window(dialog)
        if dialog.result:
            self.display_settings = self.db.get_display_settings()
            self._apply_display_settings()
            self.refresh_data()

    def _apply_display_settings(self) -> None:
        style = ttk.Style(self)
        header_font_size = int(self.display_settings.get('header_font_size', self.display_settings.get('font_size', 10)))
        data_font_size = int(self.display_settings.get('data_font_size', self.display_settings.get('font_size', 10)))
        header_row_height = int(self.display_settings.get('header_row_height', 42))
        data_row_height = int(self.display_settings.get('data_row_height', 28))
        self._tree_font = ('Microsoft JhengHei UI', data_font_size)
        self._tree_heading_font = ('Microsoft JhengHei UI', header_font_size, 'bold')
        style.configure('Hint.TLabel', font=('Microsoft JhengHei UI', FIXED_UI_FONT_SIZE))
        style.configure('Treeview', rowheight=data_row_height, font=self._tree_font, borderwidth=0, relief='flat')
        style.map('Treeview', background=[('selected', '#FFFFFF')], foreground=[('selected', '#000000')])
        style.layout('Treeview', [('Treeview.treearea', {'sticky': 'nswe'})])
        style.layout('FieldTreeview', style.layout('Treeview'))
        style.configure('FieldTreeview', rowheight=data_row_height, font=self._tree_font, borderwidth=0, relief='flat')
        style.map('FieldTreeview', background=[('selected', '#000000')], foreground=[('selected', '#FFFFFF')])
        style.layout('FieldTreeview', [('Treeview.treearea', {'sticky': 'nswe'})])
        style.configure('Treeview.Heading', font=self._tree_heading_font, padding=(6, max(4, (header_row_height - 20) // 2)))

    def refresh_fields(self) -> None:
        self.fields_tree.delete(*self.fields_tree.get_children())
        kind_names = {'text': '文字輸入', 'dropdown': '下拉選單', 'category': '分類下拉', 'computed': '自動顯示', 'lookup': '代碼下拉'}
        for field in self.db.list_fields():
            options = json.loads(field['options_json'] or '[]')
            self.fields_tree.insert('', 'end', iid=str(field['id']), values=(field['position'] + 1, field['name'], kind_names.get(field['kind'], field['kind']), ', '.join(options)))

    def refresh_lookup(self) -> None:
        self.lookup_tree.delete(*self.lookup_tree.get_children())
        for row in self.db.lookup_pairs():
            self.lookup_tree.insert('', 'end', iid=str(row['id']), values=(row['code'], row['term']))

    def import_excel(self) -> None:
        path = filedialog.askopenfilename(title='選擇 Excel 檔案', filetypes=[('Excel 檔案', '*.xlsx'), ('所有檔案', '*.*')])
        if not path:
            return
        replace = messagebox.askyesno('匯入方式', '要先清除目前資料列再重新匯入嗎？\n選擇「是」只會清除資料列，會保留欄位名稱、欄位順序、欄寬、字體及其他設定。\n選擇「否」會保留現有資料，並在後方補入 Excel 資料。', parent=self)
        try:
            fields, records = self.db.import_excel(Path(path), replace=replace, company_id=self.current_company_id)
            self.refresh_all()
            messagebox.showinfo('匯入完成', f'已處理 {fields} 個欄位、{records} 筆資料。', parent=self)
        except Exception as exc:
            messagebox.showerror('匯入失敗', str(exc), parent=self)

    def _selected_id(self, tree: ttk.Treeview) -> int | None:
        selection = tree.selection()
        if not selection:
            return None
        return int(selection[0])

    def edit_selected_record(self) -> None:
        record_id = self._selected_id(self.tree)
        if record_id is None:
            messagebox.showwarning('資料管理', '請先選取一筆資料。', parent=self)
            return
        self.open_record_editor(record_id)

    def delete_selected_record(self) -> None:
        record_id = self._selected_id(self.tree)
        if record_id is None:
            messagebox.showwarning('資料管理', '請先選取一筆資料。', parent=self)
            return
        if messagebox.askyesno('刪除資料', '確定要刪除選取的資料嗎？', parent=self):
            self.db.delete_record(record_id)
            self.refresh_data()

    def open_record_editor(self, record_id: int | None) -> None:
        win = tk.Toplevel(self)
        win.title('新增資料' if record_id is None else '編輯資料')
        win.geometry('760x760')
        win.transient(self)

        outer = ttk.Frame(win)
        outer.pack(fill='both', expand=True)
        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient='vertical', command=canvas.yview)
        form = ttk.Frame(canvas, padding=18)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')
        outer.rowconfigure(0, weight=1)
        outer.columnconfigure(0, weight=1)
        window_id = canvas.create_window((0, 0), window=form, anchor='nw')
        form.bind('<Configure>', lambda _event: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.bind('<Configure>', lambda event: canvas.itemconfigure(window_id, width=event.width))

        fields = self.db.list_fields()
        old_values = self.db.get_record_values(record_id) if record_id is not None else {}
        links = self.db.get_record_links(record_id) if record_id is not None else {}
        variables: dict[int, tk.StringVar] = {}
        link_variables: dict[int, tk.StringVar] = {}
        widgets: dict[int, tk.Widget] = {}
        field_by_key = {field['system_key']: field for field in fields if field['system_key']}
        code_field = field_by_key.get('formulation_code')
        term_field = field_by_key.get('formulation_term')

        for row, field in enumerate(fields):
            field_id = int(field['id'])
            label = field['name'].replace('\n', ' ')
            ttk.Label(form, text=label, width=26).grid(row=row, column=0, sticky='nw', padx=(0, 12), pady=4)
            var = tk.StringVar(value=old_values.get(field_id, ''))
            variables[field_id] = var
            kind = field['kind']
            if kind == 'computed':
                entry = ttk.Entry(form, textvariable=var, state='readonly', width=48)
                entry.grid(row=row, column=1, sticky='ew', pady=4)
                widgets[field_id] = entry
            elif kind in {'dropdown', 'category'}:
                options = json.loads(field['options_json'] or '[]')
                entry = ttk.Combobox(form, textvariable=var, values=options, state='readonly', width=45)
                entry.grid(row=row, column=1, sticky='ew', pady=4)
                widgets[field_id] = entry
            elif kind == 'lookup':
                options = [row['code'] for row in self.db.lookup_pairs()]
                entry = ttk.Combobox(form, textvariable=var, values=options, state='readonly', width=45)
                entry.grid(row=row, column=1, sticky='ew', pady=4)
                widgets[field_id] = entry
            else:
                entry = ttk.Entry(form, textvariable=var, width=48)
                entry.grid(row=row, column=1, sticky='ew', pady=4)
                widgets[field_id] = entry
            # 所有欄位類型都可以附加連結，包括文字、下拉選單、分類及自動顯示欄位。
            if field_id:
                link_var = tk.StringVar(value=links.get(field_id, ''))
                link_variables[field_id] = link_var
                link_buttons = ttk.Frame(form)
                link_buttons.grid(row=row, column=2, padx=(8, 0), pady=4, sticky='w')
                if link_var.get().strip():
                    ttk.Button(link_buttons, text='開啟連結', command=lambda variable=link_var: self.open_link(variable.get())).pack(side='left')
                ttk.Button(link_buttons, text='編輯連結', command=lambda variable=link_var: self.edit_link(win, variable)).pack(side='left', padx=(6, 0))

        form.columnconfigure(1, weight=1)

        def sync_term(*_args: Any) -> None:
            if not code_field or not term_field:
                return
            code = variables[int(code_field['id'])].get()
            pair = next((row for row in self.db.lookup_pairs() if row['code'] == code), None)
            variables[int(term_field['id'])].set(pair['term'] if pair else '')

        if code_field:
            variables[int(code_field['id'])].trace_add('write', sync_term)
            sync_term()

        buttons = ttk.Frame(win, padding=(18, 8, 18, 18))
        buttons.pack(fill='x')

        def save() -> None:
            values = {field_id: var.get() for field_id, var in variables.items()}
            try:
                saved_record_id = record_id if record_id is not None else self.db.create_record(values, company_id=self.current_company_id)
                if record_id is not None:
                    self.db.update_record(record_id, values)
                for field_id, link_var in link_variables.items():
                    self.db.set_record_link(saved_record_id, field_id, link_var.get())
                win.destroy()
                self.refresh_data()
            except Exception as exc:
                messagebox.showerror('儲存失敗', str(exc), parent=win)

        ttk.Button(buttons, text='取消', command=win.destroy).pack(side='right', padx=(8, 0))
        ttk.Button(buttons, text='儲存資料', command=save).pack(side='right')
        win.grab_set()

    def _record_display_name(self, record_id: int) -> str:
        fields = self.db.list_fields()
        values = self.db.get_record_values(record_id)
        brand = next((f for f in fields if f['system_key'] == 'brand_name'), None)
        permit = next((f for f in fields if f['system_key'] == 'permit_number'), None)
        brand_name = values.get(int(brand['id']), '') if brand else ''
        permit_no = values.get(int(permit['id']), '') if permit else ''
        if brand_name and permit_no:
            return f'{brand_name}（{permit_no}）'
        return brand_name or permit_no or f'資料 #{record_id}'

    def open_production_records(self, record_id: int) -> None:
        """開啟單一成品的生產履歷視窗。"""
        product_name = self._record_display_name(record_id)
        fields = self.db.list_fields()
        values = self.db.get_record_values(record_id)

        def field_value(system_key: str) -> str:
            field = next((f for f in fields if f['system_key'] == system_key), None)
            return str(values.get(int(field['id']), '') or '') if field else ''

        formulation = field_value('formulation')
        content = field_value('content')
        win = tk.Toplevel(self)
        win.title(f'生產記錄 - {product_name}')
        win.geometry('1200x780')
        win.minsize(1000, 680)
        win.transient(self)
        try:
            win.iconphoto(True, self._blank_window_icon)
        except Exception:
            pass
        try:
            style = ttk.Style(win)
            style.configure('ProductionTitle.TLabel', font=('Microsoft JhengHei UI', FIXED_UI_FONT_SIZE + 8, 'bold'))
            style.configure('ProductionSub.TLabel', font=('Microsoft JhengHei UI', FIXED_UI_FONT_SIZE + 1))
            style.configure('ProductionCard.TLabelframe', padding=12)
            style.configure('ProductionCard.TLabelframe.Label', font=('Microsoft JhengHei UI', FIXED_UI_FONT_SIZE, 'bold'))
            style.configure('ProductionCardValue.TLabel', font=('Microsoft JhengHei UI', FIXED_UI_FONT_SIZE + 10, 'bold'))
            style.configure('ProductionCardDetail.TLabel', font=('Microsoft JhengHei UI', FIXED_UI_FONT_SIZE))
            style.configure('ProductionSection.TLabel', font=('Microsoft JhengHei UI', FIXED_UI_FONT_SIZE + 3, 'bold'))
            style.configure('Production.Treeview', font=self._tree_font, rowheight=34)
            style.configure('Production.Treeview.Heading', font=self._tree_heading_font, padding=(6, max(4, (int(self.display_settings.get('header_row_height', 42)) - 20) // 2)))
        except tk.TclError:
            pass
        outer = ttk.Frame(win, padding=(18, 14, 18, 14)); outer.pack(fill='both', expand=True)
        header = ttk.Frame(outer); header.pack(fill='x', pady=(0, 14))
        title_frame = ttk.Frame(header); title_frame.pack(side='left', fill='x', expand=True)
        ttk.Label(title_frame, text=product_name, style='ProductionTitle.TLabel').pack(anchor='w')
        subtitle_parts = [part for part in (formulation, content) if part]
        latest_rows = self.db.production_records(record_id, limit=1)
        latest_date = str(latest_rows[0]['production_date'] or '') if latest_rows else ''
        if latest_date: subtitle_parts.append(f'最近生產 {latest_date}')
        ttk.Label(title_frame, text=' / '.join(subtitle_parts) if subtitle_parts else '尚無生產記錄', style='ProductionSub.TLabel').pack(anchor='w', pady=(3, 0))
        ttk.Button(header, text='新增生產記錄', command=lambda: edit_record(None)).pack(side='right', anchor='n')
        summary = ttk.Frame(outer); summary.pack(fill='x', pady=(0, 18)); summary.columnconfigure(0, weight=1); summary.columnconfigure(1, weight=1)
        def make_card(parent: tk.Misc, col: int, heading: str):
            card = ttk.LabelFrame(parent, text=heading, style='ProductionCard.TLabelframe'); card.grid(row=0, column=col, sticky='nsew', padx=(0,8) if col==0 else (8,0))
            value_label=ttk.Label(card,text='0',style='ProductionCardValue.TLabel'); value_label.pack(anchor='w',pady=(0,2))
            detail_label=ttk.Label(card,text='',style='ProductionCardDetail.TLabel'); detail_label.pack(anchor='w')
            return value_label, detail_label
        total_value,total_detail=make_card(summary,0,'記錄總數'); latest_value,latest_detail=make_card(summary,1,'最新下單數量')
        section=ttk.Frame(outer); section.pack(fill='both',expand=True)
        section_top=ttk.Frame(section); section_top.pack(fill='x',pady=(0,8))
        section_title=ttk.Label(section_top,text='生產歷程',style='ProductionSection.TLabel'); section_title.pack(side='left')
        ttk.Button(section_top,text='異常訂單',command=lambda: show_abnormal_orders()).pack(side='right',padx=(8,0))
        ttk.Button(section_top,text='管理欄位',command=lambda: manage_columns()).pack(side='right')
        table_wrap=ttk.Frame(section); table_wrap.pack(fill='both',expand=True)
        table=ttk.Treeview(table_wrap,show='headings',selectmode='browse',style='Production.Treeview')
        yscroll=ttk.Scrollbar(table_wrap,orient='vertical',command=table.yview); table.configure(yscrollcommand=yscroll.set)
        table.grid(row=0,column=0,sticky='nsew'); yscroll.grid(row=0,column=1,sticky='ns'); table_wrap.rowconfigure(0,weight=1); table_wrap.columnconfigure(0,weight=1)
        all_data_columns=('date','order','stock','manufacturer','remark')
        headings={'date':('生產日期',145,'center'),'order':('下單數量',125,'center'),'stock':('入庫數量',125,'center'),'manufacturer':('製作廠商',210,'w'),'remark':('備註',1,'w')}
        self.production_display_settings=self.db.get_display_settings(); page_size=100; page_var=tk.IntVar(value=0); page_info=tk.StringVar(value='')
        def configure_columns():
            visible=[key for key in all_data_columns if key in self.production_display_settings.get('production_visible_columns',all_data_columns)]
            table['columns']=('seq',*visible,'actions'); table.heading('seq',text='序號',anchor='center'); table.column('seq',width=60,minwidth=55,anchor='center',stretch=False)
            for key in visible:
                text,width,anchor=headings[key]; table.heading(key,text=text,anchor='center'); table.column(key,width=width,minwidth=90 if key!='remark' else 160,anchor=anchor,stretch=(key=='remark'))
            table.heading('actions',text='操作',anchor='center'); table.column('actions',width=190,minwidth=175,anchor='center',stretch=False)
        configure_columns()
        def selected_id():
            sel=table.selection(); return int(sel[0]) if sel else None
        def mismatch(order,stock):
            a=str(order or '').strip().replace(',',''); b=str(stock or '').strip().replace(',','')
            return bool(a or b) and a!=b
        def refresh():
            page=max(0,int(page_var.get())); total=self.db.production_record_count(record_id); abnormal_total=self.db.production_record_count(record_id,abnormal_only=True); max_page=max(0,(total-1)//page_size)
            if page>max_page: page=max_page; page_var.set(page)
            rows=self.db.production_records(record_id,limit=page_size,offset=page*page_size); total_value.config(text=f'{total:,}'); total_detail.config(text=f'筆生產記錄，其中異常 {abnormal_total:,} 筆')
            latest=self.db.production_records(record_id,limit=1)
            if latest:
                latest_text=str(latest[0]['order_quantity'] or latest[0]['quantity'] or '').strip(); latest_value.config(text=latest_text or '0'); latest_detail.config(text=str(latest[0]['production_date'] or ''))
            else: latest_value.config(text='0'); latest_detail.config(text='尚無資料')
            section_title.config(text=f'生產歷程 {total:,} 筆記錄'); table.delete(*table.get_children()); visible=[key for key in all_data_columns if key in self.production_display_settings.get('production_visible_columns',all_data_columns)]
            for index,row in enumerate(rows,start=page*page_size+1):
                order=str(row['order_quantity'] or row['quantity'] or ''); stock=str(row['stock_quantity'] or ''); data={'date':str(row['production_date'] or ''),'order':order,'stock':stock,'manufacturer':str(row['manufacturer'] or ''),'remark':str(row['remark'] or '')}; action='開啟連結 ／ 刪除' if str(row['external_url'] or '').strip() else '無連結 ／ 刪除'; tags=('abnormal',) if mismatch(order,stock) else ()
                table.insert('', 'end', iid=str(row['id']), values=[f'{index:02d}']+[data[key] for key in visible]+[action], tags=tags)
            table.tag_configure('abnormal',background='#FCE4E4'); page_count=max(1,max_page+1); page_info.set(f'第 {page+1} / {page_count} 頁  （每頁 {page_size} 筆）'); prev_btn.config(state='normal' if page>0 else 'disabled'); next_btn.config(state='normal' if page<max_page else 'disabled'); self.refresh_data()
        def edit_record(existing_id=None):
            old=self.db.production_record(existing_id) if existing_id is not None else None; dialog=tk.Toplevel(win); dialog.title('新增生產記錄' if old is None else '編輯生產記錄'); dialog.resizable(False,False)
            form=ttk.Frame(dialog,padding=18); form.pack(fill='both',expand=True); form.columnconfigure(1,weight=1)
            vars_={'date':tk.StringVar(value='' if old is None else old['production_date']),'order':tk.StringVar(value='' if old is None else (old['order_quantity'] or old['quantity'])),'stock':tk.StringVar(value='' if old is None else old['stock_quantity']),'manufacturer':tk.StringVar(value='' if old is None else old['manufacturer']),'remark':tk.StringVar(value='' if old is None else old['remark']),'url':tk.StringVar(value='' if old is None else old['external_url'])}
            for i,(key,label) in enumerate([('date','生產日期'),('order','下單數量'),('stock','入庫數量')]): ttk.Label(form,text=label).grid(row=i,column=0,sticky='w',padx=(0,12),pady=6); ttk.Entry(form,textvariable=vars_[key],width=42).grid(row=i,column=1,sticky='ew',pady=6)
            ttk.Label(form,text='製作廠商').grid(row=3,column=0,sticky='w',padx=(0,12),pady=6); combo=ttk.Combobox(form,textvariable=vars_['manufacturer'],width=39,state='normal',values=[r['name'] for r in self.db.production_manufacturers(self.current_company_id)]); combo.grid(row=3,column=1,sticky='ew',pady=6); ttk.Button(form,text='管理廠商',command=lambda:manage_manufacturers(dialog,combo)).grid(row=3,column=2,padx=(8,0),pady=6)
            ttk.Label(form,text='備註').grid(row=4,column=0,sticky='w',padx=(0,12),pady=6); ttk.Entry(form,textvariable=vars_['remark'],width=42).grid(row=4,column=1,sticky='ew',pady=6)
            ttk.Label(form,text='外部連結').grid(row=5,column=0,sticky='w',padx=(0,12),pady=6); ttk.Entry(form,textvariable=vars_['url'],width=42).grid(row=5,column=1,sticky='ew',pady=6); ttk.Label(form,text='例如：https://...',style='Hint.TLabel').grid(row=6,column=1,sticky='w')
            buttons=ttk.Frame(form); buttons.grid(row=7,column=0,columnspan=3,sticky='e',pady=(14,0))
            def save():
                try:
                    if not vars_['date'].get().strip(): messagebox.showwarning('生產記錄','請輸入生產日期。',parent=dialog); return
                    manufacturer=vars_['manufacturer'].get().strip(); self.db.ensure_production_manufacturer(manufacturer, self.current_company_id)
                    if old is None: self.db.create_production_record(record_id,vars_['date'].get(),vars_['order'].get(),vars_['stock'].get(),manufacturer,vars_['remark'].get(),external_url=vars_['url'].get())
                    else: self.db.update_production_record(existing_id,vars_['date'].get(),vars_['order'].get(),vars_['stock'].get(),manufacturer,vars_['remark'].get(),external_url=vars_['url'].get())
                    dialog.destroy(); refresh()
                except Exception as exc: messagebox.showerror('生產記錄',str(exc),parent=dialog)
            ttk.Button(buttons,text='取消',command=dialog.destroy).pack(side='right',padx=(8,0)); ttk.Button(buttons,text='儲存',command=save).pack(side='right'); dialog.transient(win); dialog.grab_set(); dialog.focus_set(); combo.focus_set()
        def edit_selected():
            pid=selected_id()
            if pid is None: messagebox.showwarning('生產記錄','請先選取一筆生產記錄。',parent=win); return
            edit_record(pid)
        def manage_columns():
            dialog=tk.Toplevel(win); dialog.title('管理生產履歷欄位'); dialog.resizable(False,False); body=ttk.Frame(dialog,padding=18); body.pack(fill='both',expand=True); ttk.Label(body,text='選擇要顯示的欄位',style='ProductionSection.TLabel').pack(anchor='w',pady=(0,10)); labels={'date':'生產日期','order':'下單數量','stock':'入庫數量','manufacturer':'製作廠商','remark':'備註'}; current=set(self.production_display_settings.get('production_visible_columns',all_data_columns)); vars_={}
            for key in all_data_columns: vars_[key]=tk.BooleanVar(value=key in current); ttk.Checkbutton(body,text=labels[key],variable=vars_[key]).pack(fill='x',pady=3)
            buttons=ttk.Frame(body); buttons.pack(fill='x',pady=(14,0))
            def save_columns():
                selected=[key for key in all_data_columns if vars_[key].get()]
                if not selected: messagebox.showwarning('管理欄位','至少要保留一個欄位。',parent=dialog); return
                self.production_display_settings['production_visible_columns']=selected; self.db.save_display_settings(self.production_display_settings); configure_columns(); dialog.destroy(); refresh()
            ttk.Button(buttons,text='取消',command=dialog.destroy).pack(side='right',padx=(8,0)); ttk.Button(buttons,text='儲存',command=save_columns).pack(side='right'); dialog.transient(win); dialog.grab_set()
        def show_abnormal_orders():
            dialog=tk.Toplevel(win); dialog.title(f'異常訂單 - {product_name}'); dialog.geometry('1080x600'); dialog.minsize(900,500); dialog.transient(win); body=ttk.Frame(dialog,padding=16); body.pack(fill='both',expand=True); top=ttk.Frame(body); top.pack(fill='x',pady=(0,10)); count=self.db.production_record_count(record_id,abnormal_only=True); ttk.Label(top,text=f'下單數量與入庫數量不符合，共 {count:,} 筆',style='ProductionSection.TLabel').pack(side='left')
            frame=ttk.Frame(body); frame.pack(fill='both',expand=True); cols=('seq','date','order','stock','manufacturer','remark','actions'); tree2=ttk.Treeview(frame,columns=cols,show='headings',selectmode='browse',style='Production.Treeview')
            for key,text,width,anchor in [('seq','序號',60,'center'),('date','生產日期',145,'center'),('order','下單數量',125,'center'),('stock','入庫數量',125,'center'),('manufacturer','製作廠商',190,'w'),('remark','備註',1,'w'),('actions','操作',190,'center')]: tree2.heading(key,text=text,anchor='center'); tree2.column(key,width=width,minwidth=90 if key!='remark' else 160,anchor=anchor,stretch=(key=='remark'))
            scroll=ttk.Scrollbar(frame,orient='vertical',command=tree2.yview); tree2.configure(yscrollcommand=scroll.set); tree2.grid(row=0,column=0,sticky='nsew'); scroll.grid(row=0,column=1,sticky='ns'); frame.rowconfigure(0,weight=1); frame.columnconfigure(0,weight=1)
            rows=self.db.production_records(record_id,abnormal_only=True)
            for index,row in enumerate(rows,1): tree2.insert('', 'end',iid=str(row['id']),values=(f'{index:02d}',str(row['production_date'] or ''),str(row['order_quantity'] or row['quantity'] or ''),str(row['stock_quantity'] or ''),str(row['manufacturer'] or ''),str(row['remark'] or ''),'開啟連結 ／ 刪除' if str(row['external_url'] or '').strip() else '無連結 ／ 刪除'),tags=('abnormal',))
            tree2.tag_configure('abnormal',background='#FCE4E4')
            def click_abnormal(event):
                row_id=tree2.identify_row(event.y); col=tree2.identify_column(event.x)
                if not row_id or col!='#7': return
                bbox=tree2.bbox(row_id,'actions')
                if not bbox: return
                row=self.db.production_record(int(row_id))
                if event.x-bbox[0] < bbox[2]*0.65:
                    if row and str(row['external_url'] or '').strip(): self.open_link(str(row['external_url']))
                    else: messagebox.showinfo('外部連結','這筆生產記錄沒有設定外部連結。',parent=dialog)
                elif messagebox.askyesno('刪除生產記錄','確定要刪除選取的生產記錄嗎？',parent=dialog): self.db.delete_production_record(int(row_id)); dialog.destroy(); refresh()
            tree2.bind('<Button-1>',click_abnormal,add='+'); ttk.Button(body,text='關閉',command=dialog.destroy).pack(side='right',pady=(10,0)); dialog.grab_set()
        def table_click(event):
            if table.identify('region',event.x,event.y)!='cell': return
            row_id=table.identify_row(event.y); col=table.identify_column(event.x)
            if not row_id: return
            current_columns=list(table['columns'])
            if col==f"#{current_columns.index('actions')+1}":
                bbox=table.bbox(row_id,'actions')
                if not bbox: return
                row=self.db.production_record(int(row_id)); local_x=event.x-bbox[0]
                if local_x < bbox[2]*0.65:
                    if row and str(row['external_url'] or '').strip(): self.open_link(str(row['external_url']))
                    else: messagebox.showinfo('外部連結','這筆生產記錄沒有設定外部連結。',parent=win)
                elif messagebox.askyesno('刪除生產記錄','確定要刪除選取的生產記錄嗎？',parent=win): self.db.delete_production_record(int(row_id)); refresh()
        pager=ttk.Frame(outer); pager.pack(fill='x',pady=(10,0)); prev_btn=ttk.Button(pager,text='上一頁',command=lambda:(page_var.set(max(0,page_var.get()-1)),refresh())); prev_btn.pack(side='left'); ttk.Label(pager,textvariable=page_info,style='Hint.TLabel').pack(side='left',padx=12); next_btn=ttk.Button(pager,text='下一頁',command=lambda:(page_var.set(page_var.get()+1),refresh())); next_btn.pack(side='left')
        footer=ttk.Frame(outer); footer.pack(fill='x',pady=(8,0)); ttk.Label(footer,text='點擊「開啟連結」開啟外部網址；「刪除」可刪除記錄；雙擊其他欄位可編輯。異常訂單會以淺紅色標示。',style='Hint.TLabel').pack(side='left'); ttk.Button(footer,text='關閉',command=win.destroy).pack(side='right')
        table.bind('<Button-1>',table_click,add='+'); table.bind('<Double-1>',lambda _event:edit_selected()); refresh()

    def edit_link(self, parent: tk.Misc, link_var: tk.StringVar) -> None:
        dialog = LinkDialog(parent, link_var.get())
        self.wait_window(dialog)
        if dialog.result is not None:
            link_var.set(dialog.result)

    @staticmethod
    def open_link(url: str) -> None:
        url = url.strip()
        if not url:
            return
        import os
        import webbrowser
        try:
            webbrowser.open(url)
        except Exception:
            if sys.platform.startswith('win'):
                os.startfile(url)  # type: ignore[attr-defined]

    def add_field(self) -> None:
        dialog = FieldDialog(self, self.db)
        self.wait_window(dialog)
        if dialog.result:
            self.refresh_all()

    def edit_selected_field(self) -> None:
        field_id = self._selected_id(self.fields_tree)
        if field_id is None:
            messagebox.showwarning('欄位管理', '請先選取欄位。', parent=self)
            return
        field = self.db.get_field(field_id)
        if field is None:
            return
        dialog = FieldDialog(self, self.db, field)
        self.wait_window(dialog)
        if dialog.result:
            self.refresh_all()

    def edit_category_colors(self) -> None:
        field = next((field for field in self.db.list_fields() if field['system_key'] == 'formulation'), None)
        if field is None:
            messagebox.showinfo('欄位管理', '目前沒有劑型欄位。', parent=self)
            return
        options = json.loads(field['options_json'] or '[]')
        colors = json.loads(field['color_map_json'] or '{}')
        dialog = ColorDialog(self, self.db, int(field['id']), options, colors)
        self.wait_window(dialog)
        self.refresh_all()

    def add_lookup(self) -> None:
        dialog = LookupDialog(self, '新增代碼／名稱對照')
        self.wait_window(dialog)
        if dialog.result:
            try:
                self.db.add_lookup(*dialog.result)
                self.refresh_all()
            except ValueError as exc:
                messagebox.showerror('代碼／名稱對照', str(exc), parent=self)

    def edit_selected_lookup(self) -> None:
        lookup_id = self._selected_id(self.lookup_tree)
        if lookup_id is None:
            messagebox.showwarning('代碼／名稱對照', '請先選取一筆對照資料。', parent=self)
            return
        row = next((row for row in self.db.lookup_pairs() if int(row['id']) == lookup_id), None)
        if row is None:
            return
        dialog = LookupDialog(self, '編輯代碼／名稱對照', row['code'], row['term'])
        self.wait_window(dialog)
        if dialog.result:
            try:
                self.db.update_lookup(lookup_id, *dialog.result)
                self.refresh_all()
            except ValueError as exc:
                messagebox.showerror('代碼／名稱對照', str(exc), parent=self)

    def delete_selected_lookup(self) -> None:
        lookup_id = self._selected_id(self.lookup_tree)
        if lookup_id is None:
            messagebox.showwarning('代碼／名稱對照', '請先選取一筆對照資料。', parent=self)
            return
        if messagebox.askyesno('刪除對照', '刪除後，已使用此代碼的資料仍會保留，但名稱不會再自動更新。確定刪除嗎？', parent=self):
            self.db.delete_lookup(lookup_id)
            self.refresh_all()

    def backup_now(self) -> None:
        try:
            target = self.db.backup()
            messagebox.showinfo('備份完成', f'備份檔已建立：\n{target}', parent=self)
        except Exception as exc:
            messagebox.showerror('備份失敗', str(exc), parent=self)

    def backup_as(self) -> None:
        target = filedialog.asksaveasfilename(title='另存資料庫備份', defaultextension='.db', filetypes=[('SQLite 資料庫', '*.db'), ('所有檔案', '*.*')])
        if not target:
            return
        try:
            self.db.backup(Path(target))
            messagebox.showinfo('備份完成', f'已儲存至：\n{target}', parent=self)
        except Exception as exc:
            messagebox.showerror('備份失敗', str(exc), parent=self)

    def restore_backup(self) -> None:
        source = filedialog.askopenfilename(title='選擇資料庫備份', filetypes=[('SQLite 資料庫', '*.db'), ('所有檔案', '*.*')])
        if not source:
            return
        if not messagebox.askyesno('還原資料庫', '還原會覆蓋目前資料，確定繼續嗎？', parent=self):
            return
        try:
            self.db.restore(Path(source))
            self.refresh_all()
            messagebox.showinfo('還原完成', '資料庫已還原。', parent=self)
        except Exception as exc:
            messagebox.showerror('還原失敗', str(exc), parent=self)

    def _on_close(self) -> None:
        self._save_current_column_widths()
        self.db.close()
        self.destroy()


def self_test() -> None:
    source = app_dir() / DEFAULT_XLSX
    with tempfile.TemporaryDirectory(prefix='offline_excel_db_test_') as temp_dir:
        db = Database(Path(temp_dir))
        field_count, record_count = db.import_excel(source, replace=True)
        assert field_count == 15, field_count
        assert record_count == 3, record_count
        fields = db.list_fields()
        assert len(fields) == 15
        by_name = {field['name']: field for field in fields}
        by_key = {field['system_key']: field for field in fields if field['system_key']}
        assert by_key['formulation_term']['kind'] == 'computed'
        assert db.get_record_values(1)[int(by_key['formulation_term']['id'])] == '水溶性粒劑'

        # 所有內建欄位名稱都可以改名；連動功能使用 system_key，不依賴顯示名稱。
        db.rename_field(int(by_key['approval']['id']), '核准狀態')
        db.rename_field(int(by_key['formulation']['id']), '劑型分類')
        db.rename_field(int(by_key['formulation_code']['id']), '劑型代碼')
        db.rename_field(int(by_key['formulation_term']['id']), '劑型中文名稱')
        renamed_fields = {field['system_key']: field for field in db.list_fields() if field['system_key']}
        assert db.get_field(int(renamed_fields['approval']['id']))['name'] == '核准狀態'
        custom_id = db.add_field('測試欄位', 'dropdown', ['甲', '乙'])
        new_id = db.create_record({custom_id: '乙'})
        assert db.get_record_values(new_id)[custom_id] == '乙'
        db.update_record(new_id, {int(renamed_fields['formulation_code']['id']): 'SC'})
        assert db.get_record_values(new_id)[int(renamed_fields['formulation_term']['id'])] == '水懸劑'
        backup_path = db.backup(Path(temp_dir) / 'backup.db')
        assert backup_path.exists()
        db.close()
        print(f'SELF_TEST_OK fields={field_count} records={record_count} custom_record={new_id}')


def main() -> None:
    if '--self-test' in sys.argv:
        self_test()
        return
    root = app_dir()
    db = Database(root)
    try:
        if not db.has_data():
            source = root / DEFAULT_XLSX
            if source.exists():
                db.import_excel(source, replace=False)
        app = OfflineDatabaseApp(db)
        app.mainloop()
    finally:
        try:
            db.close()
        except Exception:
            pass


if __name__ == '__main__':
    main()
